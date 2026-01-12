#!/usr/bin/env python3
"""
Excel Comparison Test for Calculation Engine

Compares calculation engine output against actual Excel reference values.
Uses test_raschet_multi_currency_correct_rate_2711_30pct_100k_29-11_vat22.xlsm

IMPORTANT: Tests multi-product calculation with distributed logistics.
Tolerance: 0.01%

Usage:
    python test_excel_comparison.py
"""

from decimal import Decimal
from datetime import date
from typing import Dict, Any, List, Tuple
from openpyxl import load_workbook

from calculation_engine import calculate_multiproduct_quote
from calculation_models import (
    QuoteCalculationInput,
    ProductInfo,
    FinancialParams,
    LogisticsParams,
    TaxesAndDuties,
    PaymentTerms,
    CustomsAndClearance,
    CompanySettings,
    SystemConfig,
    Currency,
    SupplierCountry,
    SellerCompany,
    OfferSaleType,
    Incoterms,
    DMFeeType
)


# Excel file path (has actual test data with 5 products)
REFERENCE_FILE = "/Users/andreynovikov/workspace/tech/projects/kvota/user-feedback/validation_data/test_raschet_multi_currency_correct_rate_2711_30pct_100k_29-11_vat22.xlsm"


def safe_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    """Safely convert value to Decimal."""
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value))
    except (ValueError, TypeError):
        return default


def load_excel_data(file_path: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Load quote-level settings and product data from Excel.

    Returns:
        Tuple of (quote_settings, products_list)
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb["расчет"]

    # Quote-level settings
    quote_settings = {
        # Company settings
        "seller_company": ws["D5"].value,
        "offer_sale_type": ws["D6"].value,
        "offer_incoterms": ws["D7"].value,
        "currency_of_quote": ws["D8"].value,
        "delivery_time": ws["D9"].value,

        # Payment terms
        "advance_from_client": ws["J5"].value,  # 1 = 100%
        "advance_to_supplier": ws["D11"].value,  # 1 = 100%
        "time_to_advance": ws["K5"].value or 0,
        "time_to_payment": ws["K9"].value or 10,

        # Logistics (quote-level totals)
        "logistics_supplier_hub": ws["W2"].value,
        "logistics_hub_customs": ws["W3"].value,
        "logistics_customs_client": ws["W4"].value,

        # Brokerage
        "brokerage_hub": ws["W5"].value,
        "brokerage_customs": ws["W6"].value,
        "warehousing_at_customs": ws["W7"].value,
        "customs_documentation": ws["W8"].value,
        "brokerage_extra": ws["W9"].value,

        # DM Fee
        "dm_fee_type": ws["AG3"].value,  # Фикс
        "dm_fee_value": ws["AG4"].value,  # 1000

        # Admin rates
        "rate_forex_risk": ws["AH11"].value,  # 0.03 (3%)
        "rate_fin_comm": ws["AI11"].value,  # 0.02 (2%)
    }

    # Product data (rows 16-20, 5 products)
    products = []
    for row in range(16, 21):
        product_name = ws[f"D{row}"].value
        if not product_name:
            continue

        product = {
            "row": row,
            "product_name": product_name,

            # Inputs
            "quantity": ws[f"E{row}"].value,
            "weight_in_kg": ws[f"G{row}"].value or 0,
            "currency_of_base_price": ws[f"J{row}"].value,
            "base_price_vat": ws[f"K{row}"].value,
            "supplier_country": ws[f"L{row}"].value,
            "vat_rate": ws[f"M{row}"].value,  # 0.2 = 20%
            "supplier_discount": ws[f"O{row}"].value or 0,
            "exchange_rate": ws[f"Q{row}"].value,
            "customs_code": str(ws[f"W{row}"].value) if ws[f"W{row}"].value else "0000000000",
            "import_tariff": ws[f"X{row}"].value,  # 0.05 = 5%
            "markup": ws[f"AC{row}"].value,  # 0.15 = 15%

            # Expected outputs
            "expected": {
                "N16": ws[f"N{row}"].value,
                "P16": ws[f"P{row}"].value,
                "R16": ws[f"R{row}"].value,
                "S16": ws[f"S{row}"].value,
                "T16": ws[f"T{row}"].value,
                "U16": ws[f"U{row}"].value,
                "V16": ws[f"V{row}"].value,
                "Y16": ws[f"Y{row}"].value,
                "AA16": ws[f"AA{row}"].value,
                "AB16": ws[f"AB{row}"].value,
                "AD16": ws[f"AD{row}"].value,
                "AE16": ws[f"AE{row}"].value,
                "AF16": ws[f"AF{row}"].value,
                "AG16": ws[f"AG{row}"].value,
                "AH16": ws[f"AH{row}"].value,
                "AI16": ws[f"AI{row}"].value,
                "AJ16": ws[f"AJ{row}"].value,
                "AK16": ws[f"AK{row}"].value,
                "AL16": ws[f"AL{row}"].value,
                "AN16": ws[f"AN{row}"].value,
                "AO16": ws[f"AO{row}"].value,
                "AP16": ws[f"AP{row}"].value,
                "AQ16": ws[f"AQ{row}"].value,
                "AX16": ws[f"AX{row}"].value,
                "AY16": ws[f"AY{row}"].value,
                "BA16": ws[f"BA{row}"].value,
                "BB16": ws[f"BB{row}"].value,
                "BD16": ws[f"BD{row}"].value,
            }
        }
        products.append(product)

    wb.close()
    return quote_settings, products


def map_supplier_country(country_str: str) -> SupplierCountry:
    """Map Excel country string to SupplierCountry enum."""
    country_map = {
        "Турция": SupplierCountry.TURKEY,
        "Турция (транзитная зона)": SupplierCountry.TURKEY_TRANSIT,
        "Россия": SupplierCountry.RUSSIA,
        "Китай": SupplierCountry.CHINA,
        "Литва": SupplierCountry.LITHUANIA,
        "Латвия": SupplierCountry.LATVIA,
        "Болгария": SupplierCountry.BULGARIA,
        "Польша": SupplierCountry.POLAND,
        "ЕС (между странами ЕС)": SupplierCountry.EU_CROSS_BORDER,
        "ЕС (закупка между странами ЕС)": SupplierCountry.EU_CROSS_BORDER,  # Alternative spelling
        "ОАЭ": SupplierCountry.UAE,
    }
    return country_map.get(country_str, SupplierCountry.OTHER)


def build_calculation_inputs(
    quote_settings: Dict[str, Any],
    products: List[Dict[str, Any]]
) -> List[QuoteCalculationInput]:
    """
    Build calculation engine inputs from Excel data.

    Returns list of QuoteCalculationInput - one per product.
    Each contains shared quote-level parameters + product-specific params.
    """
    # Base logistics params (country will be set per product)
    base_logistics_kwargs = {
        "offer_incoterms": Incoterms(quote_settings["offer_incoterms"]),
        "delivery_time": int(quote_settings["delivery_time"]),
        "delivery_date": date.today(),
        "logistics_supplier_hub": safe_decimal(quote_settings["logistics_supplier_hub"]),
        "logistics_hub_customs": safe_decimal(quote_settings["logistics_hub_customs"]),
        "logistics_customs_client": safe_decimal(quote_settings["logistics_customs_client"]),
    }

    customs = CustomsAndClearance(
        brokerage_hub=safe_decimal(quote_settings["brokerage_hub"]),
        brokerage_customs=safe_decimal(quote_settings["brokerage_customs"]),
        warehousing_at_customs=safe_decimal(quote_settings["warehousing_at_customs"]),
        customs_documentation=safe_decimal(quote_settings["customs_documentation"]),
        brokerage_extra=safe_decimal(quote_settings["brokerage_extra"]),
    )

    company = CompanySettings(
        seller_company=SellerCompany.MASTER_BEARING_RU,
        offer_sale_type=OfferSaleType(quote_settings["offer_sale_type"]),
    )

    # Convert rate_fin_comm from decimal (0.02) to percentage (2)
    rate_fin_comm_val = safe_decimal(quote_settings["rate_fin_comm"], Decimal("0.02"))
    rate_fin_comm_pct = rate_fin_comm_val * Decimal("100") if rate_fin_comm_val <= 1 else rate_fin_comm_val

    system = SystemConfig(
        rate_fin_comm=rate_fin_comm_pct,  # Should be percentage (2 = 2%)
        rate_loan_interest_annual=Decimal("0.25"),
        rate_insurance=Decimal("0.00047"),
        customs_logistics_pmt_due=10,
    )

    payment = PaymentTerms(
        # Convert decimal (1.0 = 100%) to percentage (100)
        advance_from_client=safe_decimal(quote_settings["advance_from_client"]) * Decimal("100"),
        advance_to_supplier=safe_decimal(quote_settings["advance_to_supplier"]) * Decimal("100"),
        time_to_advance=int(quote_settings.get("time_to_advance") or 0),
        advance_on_loading=Decimal("0"),
        time_to_advance_loading=0,
        advance_on_going_to_country_destination=Decimal("0"),
        time_to_advance_going_to_country_destination=0,
        advance_on_customs_clearance=Decimal("0"),
        time_to_advance_on_customs_clearance=0,
        time_to_advance_on_receiving=int(quote_settings.get("time_to_payment") or 10),
    )

    # Product inputs
    calc_inputs = []
    for p in products:
        # Convert percentage fields from decimal (0.15) to percentage (15)
        markup_val = safe_decimal(p["markup"])
        markup_pct = markup_val * Decimal("100") if markup_val <= 1 else markup_val

        tariff_val = safe_decimal(p["import_tariff"])
        tariff_pct = tariff_val * Decimal("100") if tariff_val <= 1 else tariff_val

        discount_val = safe_decimal(p["supplier_discount"])
        discount_pct = discount_val * Decimal("100") if discount_val <= 1 else discount_val

        forex_val = safe_decimal(quote_settings["rate_forex_risk"])
        forex_pct = forex_val * Decimal("100") if forex_val <= 1 else forex_val

        product_info = ProductInfo(
            base_price_VAT=safe_decimal(p["base_price_vat"]),
            quantity=int(p["quantity"]),
            weight_in_kg=safe_decimal(p["weight_in_kg"]),
            currency_of_base_price=Currency(p["currency_of_base_price"]),
            customs_code=p["customs_code"],
        )

        financial = FinancialParams(
            currency_of_quote=Currency(quote_settings["currency_of_quote"]),
            exchange_rate_base_price_to_quote=safe_decimal(p["exchange_rate"]),
            supplier_discount=discount_pct,
            markup=markup_pct,
            rate_forex_risk=forex_pct,
            dm_fee_type=DMFeeType.FIXED if quote_settings["dm_fee_type"] == "Фикс" else DMFeeType.PERCENTAGE,
            dm_fee_value=safe_decimal(quote_settings["dm_fee_value"]),
        )

        taxes = TaxesAndDuties(
            import_tariff=tariff_pct,
            excise_tax=Decimal("0"),
            util_fee=Decimal("0"),
        )

        # Create logistics with product-specific country
        supplier_country = map_supplier_country(p["supplier_country"])
        logistics = LogisticsParams(
            supplier_country=supplier_country,
            **base_logistics_kwargs
        )

        calc_input = QuoteCalculationInput(
            product=product_info,
            financial=financial,
            logistics=logistics,
            taxes=taxes,
            payment=payment,
            customs=customs,
            company=company,
            system=system,
        )
        calc_inputs.append(calc_input)

    return calc_inputs


def compare_results(
    products: List[Dict[str, Any]],
    results: List[Any],
    tolerance_pct: Decimal = Decimal("0.01")
) -> Tuple[int, int, List[str]]:
    """
    Compare calculation results against Excel expected values.

    Args:
        products: Product data with expected values
        results: Calculation results
        tolerance_pct: Maximum allowed percentage difference (0.01 = 0.01%)

    Returns:
        Tuple of (passed_count, failed_count, error_messages)
    """
    passed = 0
    failed = 0
    errors = []

    # Field mappings: result attribute -> Excel cell
    field_mappings = [
        ("purchase_price_no_vat", "N16", "Price without VAT"),
        ("purchase_price_after_discount", "P16", "After discount"),
        ("purchase_price_per_unit_quote_currency", "R16", "Per unit in quote currency"),
        ("purchase_price_total_quote_currency", "S16", "Total purchase price"),
        ("logistics_first_leg", "T16", "Logistics first leg"),
        ("logistics_last_leg", "U16", "Logistics last leg"),
        ("logistics_total", "V16", "Total logistics"),
        ("customs_fee", "Y16", "Customs fee"),
        ("cogs_per_unit", "AA16", "COGS per unit"),
        ("cogs_per_product", "AB16", "COGS total"),
        ("sale_price_per_unit_excl_financial", "AD16", "Sale price (excl fin)"),
        ("profit", "AF16", "Profit"),
        ("dm_fee", "AG16", "DM fee"),
        ("forex_reserve", "AH16", "Forex reserve"),
        ("financial_agent_fee", "AI16", "Financial agent fee"),
        ("sales_price_per_unit_no_vat", "AJ16", "Sales price/unit no VAT"),
        ("sales_price_total_no_vat", "AK16", "Sales price total no VAT"),
        ("sales_price_total_with_vat", "AL16", "Final price with VAT"),
        ("vat_from_sales", "AN16", "VAT from sales"),
        ("vat_on_import", "AO16", "Import VAT"),
        ("vat_net_payable", "AP16", "Net VAT payable"),
        ("transit_commission", "AQ16", "Transit commission"),
        ("internal_sale_price_per_unit", "AX16", "Internal price/unit"),
        ("internal_sale_price_total", "AY16", "Internal sale total"),
        ("financing_cost_initial", "BA16", "Financing initial"),
        ("financing_cost_credit", "BB16", "Financing credit"),
        ("distribution_base", "BD16", "Distribution base"),
    ]

    print("\n" + "=" * 120)
    print(f"{'Product':<15} {'Field':<35} {'Actual':>15} {'Expected':>15} {'Diff':>12} {'%Diff':>10} {'Result':>8}")
    print("=" * 120)

    for i, (product, result) in enumerate(zip(products, results)):
        product_name = product["product_name"][:12]
        row = product["row"]

        for attr_name, cell, description in field_mappings:
            actual = getattr(result, attr_name, None)
            expected = product["expected"].get(cell)

            if actual is None:
                actual = Decimal("0")
            if expected is None:
                expected = Decimal("0")

            actual = safe_decimal(actual)
            expected = safe_decimal(expected)

            diff = actual - expected

            # Calculate percentage difference
            if expected != 0:
                pct_diff = abs(diff / expected * 100)
            else:
                pct_diff = Decimal("0") if actual == 0 else Decimal("100")

            is_pass = pct_diff <= tolerance_pct
            status = "PASS" if is_pass else "FAIL"

            if is_pass:
                passed += 1
            else:
                failed += 1
                errors.append(f"Row {row} {cell} ({description}): actual={actual:.4f}, expected={expected:.4f}, diff={pct_diff:.4f}%")

            # Print details for all products
            print(f"{product_name:<15} {description:<35} {float(actual):>15.4f} {float(expected):>15.4f} {float(diff):>12.4f} {float(pct_diff):>9.4f}% {status:>8}")

        print("-" * 120)

    return passed, failed, errors


def run_test():
    """Run the Excel comparison test."""
    print("=" * 120)
    print("CALCULATION ENGINE TEST - Excel Comparison")
    print(f"Reference file: {REFERENCE_FILE.split('/')[-1]}")
    print("Tolerance: 0.01%")
    print("=" * 120)

    # Load Excel data
    print("\n1. Loading Excel data...")
    quote_settings, products = load_excel_data(REFERENCE_FILE)
    print(f"   Loaded {len(products)} products")
    print(f"   Quote currency: {quote_settings['currency_of_quote']}")
    print(f"   Incoterms: {quote_settings['offer_incoterms']}")
    print(f"   Logistics total: {quote_settings['logistics_supplier_hub']} + {quote_settings['logistics_hub_customs']} + {quote_settings['logistics_customs_client']}")

    # Print product summary
    print("\n   Products:")
    for p in products:
        print(f"     Row {p['row']}: {p['product_name']}, qty={p['quantity']}, price={p['base_price_vat']} {p['currency_of_base_price']}, rate={p['exchange_rate']}")

    # Build calculation inputs
    print("\n2. Building calculation inputs...")
    calc_inputs = build_calculation_inputs(quote_settings, products)
    print(f"   Created {len(calc_inputs)} calculation inputs")

    # Run calculation
    print("\n3. Running multi-product calculation...")
    try:
        results = calculate_multiproduct_quote(calc_inputs)
        print(f"   Calculation complete! Got {len(results)} results")
    except Exception as e:
        print(f"   ERROR: Calculation failed: {e}")
        import traceback
        traceback.print_exc()
        return False

    # Compare results
    print("\n4. Comparing results against Excel...")
    passed, failed, errors = compare_results(products, results, Decimal("0.01"))

    # Summary
    total = passed + failed
    pass_rate = (passed / total * 100) if total > 0 else 0

    print("\n" + "=" * 120)
    print("SUMMARY")
    print("=" * 120)
    print(f"Total comparisons: {total}")
    print(f"Passed: {passed} ({pass_rate:.2f}%)")
    print(f"Failed: {failed}")
    print(f"Tolerance: 0.01%")

    if errors:
        print("\nFailed comparisons (first 20):")
        for error in errors[:20]:
            print(f"  - {error}")
        if len(errors) > 20:
            print(f"  ... and {len(errors) - 20} more")

    return failed == 0


if __name__ == "__main__":
    success = run_test()
    exit(0 if success else 1)
