#!/usr/bin/env python3
"""Extract эталон inputs + cached output values from a calc-form .xlsm.

Track A1 of the Calculation Engine verification (see
``docs/plans/2026-05-18-calc-engine-verification-design.md`` §6/§7).

Reads one Excel "старая форма расчёта" workbook (sheet ``расчет``) with
``openpyxl(data_only=True)`` — i.e. the values Excel itself last cached —
and emits a JSON golden fixture to ``tests/golden/<name>.json``.

The golden-master test (``tests/test_calc_engine_golden_master.py``) then
reads ONLY the JSON: it feeds the extracted inputs through the production
input mapper + the locked calculation engine and asserts the engine
reproduces these cached Excel outputs within ≤0.01%.

Template variants
-----------------
All three corpus files share one template, but ``rubli_zakaz15`` is a
+1-column-shifted variant: it has an extra "Финансирование вознаграждения"
column at AH, pushing the financial-output block (forex reserve, fin-agent
fee, sale price, VAT, internal pricing, financing) one column to the right.
The extractor resolves every output cell by its **header label** in row 14
rather than a fixed letter, so both shapes are handled by one code path.

Usage
-----
    # all corpus files
    .venv/bin/python scripts/refresh_golden.py

    # a single file
    .venv/bin/python scripts/refresh_golden.py idemitsu.xlsm

Run manually whenever the эталон corpus changes — the PR diff of the JSON
fixtures then shows exactly which эталон numbers moved.
"""

from __future__ import annotations

import json
import os
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Corpus — source .xlsm → golden JSON name, plus the quote IDN from the design
# doc §4 (Testing 2 rows 51-53). The IDN is metadata only; A1 is autonomous.
# ---------------------------------------------------------------------------

CORPUS = {
    "idemitsu.xlsm": "Q-202605-0011",
    "rubli_zakaz15.xlsm": "Q-202605-0012",
    "forma_nds22_18.xlsm": "Q-202605-0014",
}

SHEET = "расчет"
HEADER_ROW = 14  # row with human column labels ("Кол-во", "Профит чистый", ...)
FIRST_PRODUCT_ROW = 16

# ---------------------------------------------------------------------------
# Input columns — IDENTICAL across all corpus files (verified 2026-05-18).
# Columns C..AC never shift; only the financial-output block does.
# ---------------------------------------------------------------------------

INPUT_COLS = {
    "article": "C",            # Артикул №
    "name": "D",               # Номенклатурное наименование (blank in forma)
    "quantity": "E",           # Кол-во
    "weight_unit": "G",        # Вес шт
    "weight_total": "I",       # Общий вес
    "currency_of_base_price": "J",   # Валюта закупки
    "price_vat": "K",          # Цена закупки в Ориг. с НДС (often blank)
    "supplier_country": "L",   # Страна Закупки
    "vat_rate": "M",           # Ставка НДС в закупке %
    "price_no_vat": "N",       # Цена закупки в Ориг. без НДС
    "supplier_discount": "O",  # Скидка поставщика
    "fx_rate": "Q",            # Курс валют к валюте КП
    "customs_code": "W",       # Код ТН ВЭД
    "import_tariff": "X",      # Пошлина в %
    "markup": "AC",            # Наценка Общая в %
}

# Output cells whose column is FIXED (purchase/logistics/customs/COGS block —
# columns N..AF never shift between templates).
FIXED_OUTPUT_COLS = {
    "N16": "N",   # purchase price no VAT
    "P16": "P",   # after discount
    "R16": "R",   # per unit in quote currency
    "S16": "S",   # total purchase price
    "T16": "T",   # logistics first leg
    "U16": "U",   # logistics last leg
    "V16": "V",   # logistics total
    "Y16": "Y",   # customs duty
    "Z16": "Z",   # excise tax
    "AA16": "AA",  # COGS per unit
    "AB16": "AB",  # COGS per product
    "AD16": "AD",  # sale price/unit excl financial
    "AE16": "AE",  # sale price total excl financial
    "AF16": "AF",  # profit
    "AG16": "AG",  # DM fee (ЛПР)
}

# Output cells resolved by HEADER LABEL — these shift +1 in the rubli variant.
# Logical cell name → header-label prefix in row 14.
LABEL_OUTPUT_COLS = {
    "AH16": "Резерв на отриц",          # forex reserve
    "AI16": "Комиссия фин/агента",      # financial agent fee
    "AJ16": "Цена продажи без НДС",     # sale price/unit no VAT
    "AK16": "Сумма продажи без НДС",    # sale price total no VAT
    "AL16": "Итого Сумма Продажи",      # sale price total with VAT
    "AM16": "Цена продажи c НДС",       # sale price/unit with VAT
    "AN16": "НДС с продаж",             # VAT from sales
    "AO16": "НДС (при импорте",         # VAT on import
    "AP16": "НДС к уплате",             # net VAT payable
    "AQ16": "Комиссия за Транзит",      # transit commission
    "AY16": "Сумма валютной специфи",   # internal sale total
    "BA16": "Итого Стоимость финанс",   # financing cost initial
    "BB16": "Комиссия за рассрочку",    # financing cost credit
}

# Quote-level total cells (row 13). The financial totals shift with the same
# +1 rule; resolved against row-14 labels too.
FIXED_TOTAL_COLS = {
    "S13": "S",    # total purchase price
    "V13": "V",    # total logistics
    "Y13": "Y",    # total customs
    "AB13": "AB",  # total COGS
    "AF13": "AF",  # total profit
}
LABEL_TOTAL_COLS = {
    "AK13": "Сумма продажи без НДС",
    "AL13": "Итого Сумма Продажи",
}

# Checkpoint cell lists — referenced by the TDD test for schema completeness
# and by the golden-master test for the assert loop. Order = calc-phase order.
PRODUCT_CHECKPOINT_CELLS = (
    list(FIXED_OUTPUT_COLS.keys()) + list(LABEL_OUTPUT_COLS.keys())
)
TOTAL_CHECKPOINT_CELLS = (
    list(FIXED_TOTAL_COLS.keys()) + list(LABEL_TOTAL_COLS.keys())
)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _num(value):
    """Coerce an Excel cell value to float, or None.

    Excel sometimes leaves a formula cell as the literal string "введи курс"
    (placeholder for an unfilled rate) — that and any other non-numeric
    string collapse to None so downstream code treats the cell as empty.
    """
    if value is None or value == "":
        return None
    if isinstance(value, bool):  # guard: bool is a subclass of int
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _build_label_index(ws) -> dict:
    """Map header row 14 label → column letter, for label-based resolution."""
    index = {}
    for col in range(3, ws.max_column + 1):
        label = ws.cell(row=HEADER_ROW, column=col).value
        if label:
            index[str(label)] = get_column_letter(col)
    return index


def _resolve_label_col(label_index: dict, label_prefix: str) -> str:
    """Find the column whose row-14 label starts with ``label_prefix``."""
    for label, col in label_index.items():
        if label.startswith(label_prefix):
            return col
    raise KeyError(f"no row-{HEADER_ROW} header label starts with {label_prefix!r}")


def _detect_template_variant(label_index: dict) -> str:
    """'shifted' if the workbook has the extra financing column, else 'standard'.

    The shifted (rubli_zakaz15) variant carries a "Финансирование
    вознаграждения" label in row 14 that the standard template lacks.
    """
    for label in label_index:
        if label.startswith("Финансирование"):
            return "shifted"
    return "standard"


# ---------------------------------------------------------------------------
# Quote-level inputs
# ---------------------------------------------------------------------------


def _extract_quote_variables(ws) -> dict:
    """Pull quote-level settings — the ``variables`` dict the shim consumes.

    Layout (verified across all 3 эталон files):
      D5  seller company        D6  offer_sale_type   D7  offer_incoterms
      D8  quote currency        D9  delivery time     D11 advance to supplier
      J5  advance from client (decimal fraction)
      K9  post-payment days     AH11/AI11 rate_forex_risk
      W2..W11 / V2..V11  the logistics-cost legend (label in V, value in W)

    Logistics & brokerage are read from the V/W legend by Russian label so
    the same code works regardless of which legend rows are populated. The
    Excel logistics-distribution formula sources first/second leg from the
    pre-summed legend totals (V11 = first-leg sum, W11 = second-leg sum):
        T16 = BD16 * V11    U16 = BD16 * W11

    IMPORTANT — insurance: ``V11`` *includes* the "Страховка" legend line
    (W10). The calculation engine, by contrast, computes insurance itself
    (``AY13 × rate_insurance``, ceil to 1dp) and adds it onto T16 in
    Phase 3. So the engine must be fed the first leg WITHOUT insurance;
    feeding it V11 directly would double-count. We therefore expose both
    ``logistics_first_leg_total`` (V11, insurance included — audit value)
    and ``logistics_first_leg_excl_insurance`` (V11 − W10) which the shim
    actually feeds the engine. ``insurance_line`` is the эталон's own W10.
    """
    # Legend: V column = label, W column = value, rows 2..11.
    legend = {}
    for row in range(2, 12):
        label = ws.cell(row=row, column=22).value  # column V
        value = _num(ws.cell(row=row, column=23).value)  # column W
        if label is not None:
            legend[str(label).strip()] = value

    quote_currency = ws["D8"].value or "USD"

    first_leg_total = _num(ws["V11"].value) or 0.0
    second_leg_total = _num(ws["W11"].value) or 0.0
    # The "Страховка (% от валютной спеки)" legend line — folded into V11.
    insurance_line = 0.0
    for label, value in legend.items():
        if label.startswith("Страховка"):
            insurance_line = value or 0.0
            break

    dm_fee = _extract_dm_fee(ws)

    return {
        "seller_company": ws["D5"].value,
        "offer_sale_type": ws["D6"].value,
        "offer_incoterms": ws["D7"].value,
        "currency_of_quote": quote_currency,
        "delivery_time": int(_num(ws["D9"].value) or 0),
        # Payment terms — Excel stores percentages as decimal fractions.
        "advance_from_client": _num(ws["J5"].value) or 0.0,
        "advance_to_supplier": _num(ws["D11"].value) or 0.0,
        "time_to_advance": int(_num(ws["K5"].value) or 0),
        "time_to_advance_on_receiving": int(_num(ws["K9"].value) or 0),
        # Admin rate — AH11 in the standard template, AI11 in the shifted one.
        "rate_forex_risk": _num(ws["AH11"].value)
        if _num(ws["AH11"].value) is not None
        else _num(ws["AI11"].value),
        # DM-fee (ЛПР) reward config — resolved from the AG3/AF4:AG7 legend.
        "dm_fee_type": dm_fee["type"],
        "dm_fee_value": dm_fee["value"],
        # The pre-summed legend leg totals the Excel T16/U16 formulas use.
        "logistics_first_leg_total": first_leg_total,
        "logistics_second_leg_total": second_leg_total,
        # First leg with the insurance line removed — what the shim feeds the
        # engine, since the engine re-adds insurance in Phase 3.
        "logistics_first_leg_excl_insurance": first_leg_total - insurance_line,
        "insurance_line": insurance_line,
        # Raw legend (audit trail; lets a reviewer see the leg breakdown).
        "legend": legend,
    }


def _extract_dm_fee(ws) -> dict:
    """Resolve the DM-fee (ЛПР reward) config from the AG3 / AF4:AG7 legend.

    Excel: ``AG16 = BD16 × VLOOKUP(AG3, AF4:AG7, 2, FALSE)``. The legend is
        AF4 "Фикс"        AG4 = fixed amount
        AF6 "%"           AG6 = percentage (decimal fraction)
        AF7 "комиссия %"  AG7 = AG6 × AB13   (the quote-level total)
    AG3 picks one. The engine takes a ``dm_fee_type`` ∈ {fixed, %} and a
    ``dm_fee_value``; this maps the Excel legend onto that pair.

    Returns ``{"type": "fixed"|"%", "value": <number>}``. When the chosen
    legend cell is blank the fee is zero (type "fixed", value 0).
    """
    key = ws["AG3"].value
    fixed_amount = _num(ws["AG4"].value) or 0.0
    pct_fraction = _num(ws["AG6"].value) or 0.0

    if key == "Фикс":
        return {"type": "fixed", "value": fixed_amount}
    if key in ("%", "комиссия %"):
        # Percentage fee. AG6 holds the rate as a fraction; engine wants
        # percent. When AG6 is blank the effective fee is 0.
        return {"type": "%", "value": pct_fraction * 100.0}
    # Unknown / blank key → no fee.
    return {"type": "fixed", "value": 0.0}


# ---------------------------------------------------------------------------
# Per-product inputs
# ---------------------------------------------------------------------------


def _is_product_row(ws, row: int) -> bool:
    """True if ``row`` holds a real product line.

    A row counts as a product when it has an article (col C) OR a name
    (col D) OR a positive quantity (col E). The first non-product row marks
    the end of the contiguous product block. This matters because
    ``forma_nds22_18`` rows have NO column-D name (only an article in C),
    and ``idemitsu`` row 17/18 are residual template rows with col E empty.
    """
    article = ws.cell(row=row, column=3).value  # C
    name = ws.cell(row=row, column=4).value  # D
    qty = _num(ws.cell(row=row, column=5).value)  # E
    has_article = article is not None and str(article).strip() != ""
    has_name = name is not None and str(name).strip() != ""
    has_qty = qty is not None and qty > 0
    return has_article or has_name or has_qty


def _extract_products(ws) -> list[dict]:
    """Extract per-product input rows (contiguous block from row 16)."""
    products = []
    row = FIRST_PRODUCT_ROW
    while row <= ws.max_row:
        if not _is_product_row(ws, row):
            break
        products.append(
            {
                "row": row,
                "article": ws[f"{INPUT_COLS['article']}{row}"].value,
                "name": ws[f"{INPUT_COLS['name']}{row}"].value,
                "quantity": int(_num(ws[f"{INPUT_COLS['quantity']}{row}"].value) or 0),
                "weight_in_kg": _num(ws[f"{INPUT_COLS['weight_unit']}{row}"].value) or 0.0,
                "currency_of_base_price": ws[f"{INPUT_COLS['currency_of_base_price']}{row}"].value,
                # K = price WITH vat (raw input). Often blank — user types N.
                "price_vat": _num(ws[f"{INPUT_COLS['price_vat']}{row}"].value),
                # N = price WITHOUT vat (the эталон-cached purchase price).
                "price_no_vat": _num(ws[f"{INPUT_COLS['price_no_vat']}{row}"].value) or 0.0,
                "supplier_country": ws[f"{INPUT_COLS['supplier_country']}{row}"].value,
                "vat_rate": _num(ws[f"{INPUT_COLS['vat_rate']}{row}"].value) or 0.0,
                "supplier_discount": _num(ws[f"{INPUT_COLS['supplier_discount']}{row}"].value) or 0.0,
                "fx_rate": _num(ws[f"{INPUT_COLS['fx_rate']}{row}"].value),
                "customs_code": ws[f"{INPUT_COLS['customs_code']}{row}"].value,
                "import_tariff": _num(ws[f"{INPUT_COLS['import_tariff']}{row}"].value) or 0.0,
                "markup": _num(ws[f"{INPUT_COLS['markup']}{row}"].value) or 0.0,
            }
        )
        row += 1
    return products


# ---------------------------------------------------------------------------
# Per-product expected outputs
# ---------------------------------------------------------------------------


def _extract_product_expected(ws, row: int, label_index: dict, currency: str) -> dict:
    """Read all эталон output cells for one product row.

    Every monetary cell is stamped with the quote currency (per design §5 —
    "каждое денежное значение помечено валютой"). Fixed-column cells use
    their letter directly; label-resolved cells use the row-14 header so the
    +1 shift of the rubli variant is absorbed transparently.
    """
    out = {}
    for name, col in FIXED_OUTPUT_COLS.items():
        out[name] = {"value": _num(ws[f"{col}{row}"].value), "currency": currency}
    for name, label in LABEL_OUTPUT_COLS.items():
        col = _resolve_label_col(label_index, label)
        out[name] = {"value": _num(ws[f"{col}{row}"].value), "currency": currency}
    return out


def _extract_totals(ws, label_index: dict, currency: str) -> dict:
    """Read row-13 quote-level total cells (S13/AB13/AF13/AK13/AL13/...)."""
    totals = {}
    for name, col in FIXED_TOTAL_COLS.items():
        totals[name] = {"value": _num(ws[f"{col}13"].value), "currency": currency}
    for name, label in LABEL_TOTAL_COLS.items():
        col = _resolve_label_col(label_index, label)
        totals[name] = {"value": _num(ws[f"{col}13"].value), "currency": currency}
    return totals


# ---------------------------------------------------------------------------
# Top-level extractor
# ---------------------------------------------------------------------------


def extract_golden(xlsm_path: str, source_name: str) -> dict:
    """Extract a complete golden fixture dict from one .xlsm file.

    Args:
        xlsm_path: Absolute path to the .xlsm эталон file.
        source_name: ASCII file name recorded in the fixture (e.g.
            ``idemitsu.xlsm``) — used to look up the quote IDN.

    Returns:
        A JSON-serialisable dict with keys ``source_xlsm``, ``quote_idn``,
        ``quote_currency``, ``template_variant``, ``inputs`` and
        ``expected`` (see design §7).

    Raises:
        RuntimeError: if the sheet's formula cells were never cached by
            Excel (``data_only`` returns ``None`` everywhere) — such a file
            cannot serve as an эталон.
    """
    wb = load_workbook(xlsm_path, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"{source_name}: sheet '{SHEET}' not found")
    ws = wb[SHEET]

    label_index = _build_label_index(ws)
    template_variant = _detect_template_variant(label_index)

    variables = _extract_quote_variables(ws)
    products = _extract_products(ws)

    if not products:
        wb.close()
        raise RuntimeError(f"{source_name}: no product rows found from row {FIRST_PRODUCT_ROW}")

    # Guard: a file whose formulas were never recalculated by Excel returns
    # None for every computed cell. Such a file cannot be an эталон.
    s13 = _num(ws["S13"].value)
    if s13 is None:
        wb.close()
        raise RuntimeError(
            f"{source_name}: S13 total is empty — workbook formulas were never "
            f"cached by Excel; cannot serve as эталон"
        )

    currency = variables["currency_of_quote"]

    expected_products = []
    for p in products:
        exp = _extract_product_expected(ws, p["row"], label_index, currency)
        exp["row"] = p["row"]
        expected_products.append(exp)

    totals = _extract_totals(ws, label_index, currency)

    wb.close()

    return {
        "source_xlsm": source_name,
        "quote_idn": CORPUS.get(source_name, ""),
        "quote_currency": currency,
        "template_variant": template_variant,
        "inputs": {
            "variables": variables,
            "products": products,
        },
        "expected": {
            "products": expected_products,
            "totals": totals,
        },
    }


def refresh_one(source_name: str, sources_dir: str, out_dir: str) -> str:
    """Extract one .xlsm → write ``<name>.json``; return the JSON path."""
    xlsm_path = os.path.join(sources_dir, source_name)
    golden = extract_golden(xlsm_path, source_name)
    json_name = os.path.splitext(source_name)[0] + ".json"
    json_path = os.path.join(out_dir, json_name)
    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump(golden, fh, ensure_ascii=False, indent=2, sort_keys=True)
    return json_path


def main(argv: list[str]) -> int:
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sources_dir = os.path.join(repo_root, "tests", "golden", "sources")
    out_dir = os.path.join(repo_root, "tests", "golden")

    targets = argv[1:] if len(argv) > 1 else list(CORPUS.keys())
    rc = 0
    for source_name in targets:
        try:
            json_path = refresh_one(source_name, sources_dir, out_dir)
            golden = json.load(open(json_path, encoding="utf-8"))
            print(
                f"OK  {source_name:22} -> {os.path.basename(json_path):22} "
                f"({len(golden['inputs']['products'])} products, "
                f"{golden['template_variant']} template)"
            )
        except Exception as exc:  # noqa: BLE001 — CLI: report and continue
            print(f"FAIL {source_name}: {exc}")
            rc = 1
    return rc


if __name__ == "__main__":
    sys.exit(main(sys.argv))
