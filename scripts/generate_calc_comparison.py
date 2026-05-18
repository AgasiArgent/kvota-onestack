#!/usr/bin/env python3
"""Track B — offline engine-vs-эталон comparison .xlsx generator.

Design: ``docs/plans/2026-05-18-calc-engine-verification-design.md`` §8.

Track A1 (``tests/test_calc_engine_golden_master.py``) proves the
Calculation Engine reproduces the «старая форма расчёта» Excel within
≤0.01% — but only as a CI test. Track B turns that same comparison into a
human-readable .xlsx artifact a non-coding reviewer can simply open.

For each of the four эталон quotes this script writes one .xlsx with:

  * a **verdict banner** — «ВСЁ СХОДИТСЯ — макс. откл. X%» or
    «РАСХОЖДЕНИЕ — N ячеек вне допуска, макс. X%»;
  * sheet **«Сравнение»** — one row per position × output cell:
    ``позиция | ячейка | поле | engine | эталон (.xlsm) | абс.откл |
    %откл | вердикт``, colour-coded (FAIL red, PASS green, expected
    per-segment difference grey);
  * sheet **«Входные данные»** — the visible inputs layer: per position
    кол-во / цена / валюта / страна / группировка, so a reader sees what
    went IN.

Reused infrastructure
---------------------
All the comparison computation — running the engine through the SAME
production path Track A1 uses (``golden_support`` shim →
``build_calculation_inputs`` → ``calculate_multiproduct_quote``, with FX /
tariff / admin-rate pinned to the эталон), the ≤0.01% tolerance, the
checkpoint maps and the segment-rule classification (design §5) — lives in
``tests.golden_support`` and is shared with Track A1. This script only
*renders* the ``ComparisonResult`` it gets back.

Segment rule (design §5)
------------------------
Kvotaflow splits logistics into a variable number of segments; the эталон
Excel always uses three legs. The per-position logistics (T16/U16/V16) and
customs (Y16/Z16) rows — and rubli's downstream per-product cells — are
therefore marked «ожидаемое посегментное расхождение», NOT FAIL. Logistics
and customs get a PASS/FAIL verdict ONLY at the quote-total level (V13/Y13).

Эталон column = the .xlsm's own cached values (from ``tests/golden/*.json``);
engine column = the production engine run. Tolerance ≤0.01%, same as A1.

Usage
-----
    # all four corpus quotes
    .venv/bin/python scripts/generate_calc_comparison.py

    # a single quote
    .venv/bin/python scripts/generate_calc_comparison.py idemitsu.json

Output: ``tmp/calc-comparison/<name>_comparison.xlsx`` — artifacts for the
user to open. The directory is created if missing.
"""

from __future__ import annotations

import os
import sys
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# The script sits in scripts/; add the repo root so ``tests.golden_support``
# imports the same way the test suite does.
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from tests import golden_support  # noqa: E402 — needs the sys.path tweak above

# ---------------------------------------------------------------------------
# Corpus — golden JSON fixtures, in design §4 order.
# ---------------------------------------------------------------------------

CORPUS = [
    "idemitsu.json",
    "rubli_zakaz15.json",
    "forma_nds22_18.json",
    "amtel_cofly.json",
]

_OUT_DIR = os.path.join(_REPO_ROOT, "tmp", "calc-comparison")


# ---------------------------------------------------------------------------
# Styling — minimal, self-contained (no dependency on the heavy legacy
# export_validation_service template machinery).
# ---------------------------------------------------------------------------

_FILL_FAIL = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
_FILL_PASS = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
_FILL_SEGMENT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

_FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
_FONT_BANNER_OK = Font(bold=True, size=14, color="38761D")
_FONT_BANNER_FAIL = Font(bold=True, size=14, color="990000")
_FONT_LABEL = Font(bold=True, size=11)

_BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

_VERDICT_FILL = {
    golden_support.VERDICT_FAIL: _FILL_FAIL,
    golden_support.VERDICT_PASS: _FILL_PASS,
    golden_support.VERDICT_SEGMENT: _FILL_SEGMENT,
}


def _num(value: Decimal) -> float:
    """Decimal → float for an Excel numeric cell."""
    return float(value)


def _write_header_row(ws: Worksheet, row: int, labels: list[str]) -> None:
    """Write one styled header row."""
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER_THIN


# ---------------------------------------------------------------------------
# Sheet «Сравнение»
# ---------------------------------------------------------------------------


def _build_comparison_sheet(
    ws: Worksheet, result: golden_support.ComparisonResult
) -> None:
    """Write the «Сравнение» sheet: banner + per-cell comparison table."""
    # --- Verdict banner -----------------------------------------------------
    banner_cell = ws.cell(row=1, column=1, value=result.banner)
    is_fail = "РАСХОЖДЕНИЕ" in result.banner
    banner_cell.font = _FONT_BANNER_FAIL if is_fail else _FONT_BANNER_OK
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 26

    # Context line — which эталон, which quote, the tolerance.
    ctx = (
        f"Эталон: {result.source_xlsm}  ·  Квота: {result.quote_idn}  ·  "
        f"Валюта КП: {result.quote_currency}  ·  Допуск: ≤0.01%  ·  "
        f"engine = Calculation Engine OneStack, эталон = кэш-значения .xlsm"
    )
    ctx_cell = ws.cell(row=2, column=1, value=ctx)
    ctx_cell.font = Font(size=10, italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Legend for the colour coding.
    legend = (
        "Цвет строки: красный — FAIL (вне допуска); зелёный — PASS; "
        "серо-жёлтый — ожидаемое посегментное расхождение (логистика/таможня "
        "по позициям — см. правило сегментов, вердикт только по итогу)."
    )
    legend_cell = ws.cell(row=3, column=1, value=legend)
    legend_cell.font = Font(size=9, italic=True, color="888888")
    legend_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
    ws.row_dimensions[3].height = 30

    # --- Table header -------------------------------------------------------
    header_row = 5
    headers = [
        "позиция", "ячейка", "поле", "engine", "эталон (.xlsm)",
        "абс.откл", "%откл", "вердикт",
    ]
    _write_header_row(ws, header_row, headers)

    # --- Data rows ----------------------------------------------------------
    row = header_row + 1
    for cmp_row in result.rows:
        fill = _VERDICT_FILL.get(cmp_row.verdict)
        values = [
            cmp_row.position,
            cmp_row.cell,
            cmp_row.field,
            _num(cmp_row.engine_val),
            _num(cmp_row.etalon_val),
            _num(cmp_row.abs_delta),
            _num(cmp_row.deviation_pct),
            cmp_row.verdict,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _BORDER_THIN
            if fill is not None:
                cell.fill = fill
            if col in (4, 5, 6):  # engine / эталон / абс.откл
                cell.number_format = "#,##0.0000"
            elif col == 7:  # %откл
                cell.number_format = '0.0000"%"'
        # A currency mismatch is the loudest possible failure — annotate it.
        # Written via the ``value=`` kwarg (not the ``.value`` setter) so the
        # overwrite is type-clean — column 8 is a plain data cell, never merged.
        if cmp_row.currency_mismatch:
            ws.cell(
                row=row,
                column=8,
                value=f"{cmp_row.verdict} — НЕСОВПАДЕНИЕ ВАЛЮТЫ",
            )
        row += 1

    # --- Column widths ------------------------------------------------------
    widths = [10, 9, 38, 16, 18, 14, 11, 34]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze the header so the table stays navigable while scrolling.
    # freeze_panes takes a cell-coordinate string — type-clean, no Cell object.
    ws.freeze_panes = f"A{header_row + 1}"


# ---------------------------------------------------------------------------
# Sheet «Входные данные»
# ---------------------------------------------------------------------------


def _build_inputs_sheet(
    ws: Worksheet, result: golden_support.ComparisonResult
) -> None:
    """Write the «Входные данные» sheet: the visible inputs layer per position.

    Makes the surface explicit (design §9): a reviewer sees exactly what
    went IN — кол-во / цена / валюта / страна / группировка — so a
    divergence can be localised to input / engine / output.
    """
    title = ws.cell(
        row=1, column=1,
        value=f"Входные данные — {result.source_xlsm} ({result.quote_idn})",
    )
    title.font = _FONT_LABEL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    note = ws.cell(
        row=2, column=1,
        value=(
            "Видимый слой входов из эталонного .xlsm — что было подано в "
            "расчёт. Цена — закупочная без НДС в валюте позиции."
        ),
    )
    note.font = Font(size=9, italic=True, color="888888")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    header_row = 4
    headers = [
        "позиция", "наименование", "кол-во", "цена (закуп., без НДС)",
        "валюта", "страна", "группировка (ТН ВЭД)",
    ]
    _write_header_row(ws, header_row, headers)

    row = header_row + 1
    for inp in result.input_rows:
        values = [
            inp.position,
            inp.product_name,
            inp.quantity,
            _num(inp.price),
            inp.currency,
            inp.country,
            inp.grouping,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _BORDER_THIN
            if col == 4:
                cell.number_format = "#,##0.0000"
        row += 1

    widths = [10, 44, 12, 22, 10, 30, 22]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    # freeze_panes takes a cell-coordinate string — type-clean, no Cell object.
    ws.freeze_panes = f"A{header_row + 1}"


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------


def generate_one(json_name: str, out_dir: str) -> tuple[str, str]:
    """Build one comparison .xlsx for one golden fixture.

    Args:
        json_name: golden fixture file name (e.g. ``idemitsu.json``).
        out_dir: directory the .xlsx is written to.

    Returns:
        ``(xlsx_path, banner_text)``.
    """
    result = golden_support.build_comparison(json_name)

    wb = Workbook()
    # A fresh Workbook() always has exactly one sheet — assert narrows the
    # Optional that openpyxl's stubs declare for wb.active.
    ws_cmp = wb.active
    assert isinstance(ws_cmp, Worksheet)
    ws_cmp.title = "Сравнение"
    _build_comparison_sheet(ws_cmp, result)

    ws_inputs = wb.create_sheet("Входные данные")
    _build_inputs_sheet(ws_inputs, result)

    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(json_name)[0]
    xlsx_path = os.path.join(out_dir, f"{base}_comparison.xlsx")
    wb.save(xlsx_path)
    return xlsx_path, result.banner


def main(argv: list[str]) -> int:
    targets = argv[1:] if len(argv) > 1 else CORPUS
    rc = 0
    for json_name in targets:
        try:
            xlsx_path, banner = generate_one(json_name, _OUT_DIR)
            rel = os.path.relpath(xlsx_path, _REPO_ROOT)
            print(f"OK  {json_name:22} -> {rel}")
            print(f"    Вердикт: {banner}")
        except Exception as exc:  # noqa: BLE001 — CLI: report and continue
            print(f"FAIL {json_name}: {exc}")
            rc = 1
    return rc


if __name__ == "__main__":
    sys.exit(main(sys.argv))
