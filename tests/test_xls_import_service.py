"""
Tests for XLS Import Service (Testing 2 row 70).

The service mirrors the «Скачать XLS» template column layout — clients edit
the downloaded XLS and re-upload it; the service matches rows back to the
КПП (invoice_items) by `idn_sku` (артикул from МОП, joined via
invoice_item_coverage → quote_items) and updates the supplier-side fields
on the matched invoice_items.

Edge cases (locked decisions, docs/plans/2026-05-25-batch-24c-decisions.md):
  - new article in XLS, not in КПП       → skip + return in `skipped`
  - article in КПП, missing from XLS     → preserve existing values (no-op)
  - duplicate article in XLS             → reject with DUPLICATES error
  - all template fields update            → price, qty, MOQ, weight, dims, …

Tests use mocked Supabase clients — same pattern as
test_xls_export_service.py — to keep the suite hermetic.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook


# ---------------------------------------------------------------------------
# Fixture: build an in-memory xlsx that matches the «Скачать XLS» template
# ---------------------------------------------------------------------------

# Column layout (RU) — mirror of services.xls_export_service.COLUMNS_RU.
HEADERS_RU = [
    "Бренд",
    "Арт. запрошенный",
    "Арт. производителя",
    "Наименование производителя",
    "Наименование",
    "Кол-во",
    "Мин. заказ",
    "Цена",
    "Срок (к.д.)",
    "Вес (кг)",
    "Размеры (В×Ш×Д мм)",
    "Примечание",
    "Покрывает",
]


def make_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Build a minimal xlsx with the template headers + the given rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Items"
    for col_idx, header in enumerate(HEADERS_RU, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get("brand"))
        ws.cell(row=row_idx, column=2, value=row.get("idn_sku"))
        ws.cell(row=row_idx, column=3, value=row.get("supplier_sku"))
        ws.cell(row=row_idx, column=4, value=row.get("manufacturer_product_name"))
        ws.cell(row=row_idx, column=5, value=row.get("product_name"))
        ws.cell(row=row_idx, column=6, value=row.get("quantity"))
        ws.cell(row=row_idx, column=7, value=row.get("minimum_order_quantity"))
        ws.cell(row=row_idx, column=8, value=row.get("purchase_price_original"))
        ws.cell(row=row_idx, column=9, value=row.get("production_time_days"))
        ws.cell(row=row_idx, column=10, value=row.get("weight_in_kg"))
        ws.cell(row=row_idx, column=11, value=row.get("dimensions"))
        ws.cell(row=row_idx, column=12, value=row.get("supplier_notes"))
        # column 13 ("Покрывает") is read-only metadata — ignored on import
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Fixture: build a mock supabase that returns the КПП state we control
# ---------------------------------------------------------------------------


def _build_mock_supabase(
    invoice_items: list[dict[str, Any]],
    updates_capture: list[dict[str, Any]] | None = None,
) -> MagicMock:
    """Mock supabase with two surfaces the service touches:

    1. ``.table("invoice_items").select(...).eq("invoice_id", id).order(...)``
       — returns the invoice_items + embedded coverage join, used to build
       the (idn_sku → invoice_item) lookup table.
    2. ``.table("invoice_items").update({...}).eq("id", iid).execute()``
       — captured into ``updates_capture`` for assertions.
    """
    captured = updates_capture if updates_capture is not None else []
    mock_sb = MagicMock()

    def table_side_effect(name: str) -> MagicMock:
        if name != "invoice_items":
            raise AssertionError(f"unexpected table access: {name}")

        # ---- SELECT branch -------------------------------------------------
        select_mock = MagicMock()
        select_eq = MagicMock()
        select_order = MagicMock()
        select_order_resp = MagicMock()
        select_order_resp.data = invoice_items
        select_order.execute.return_value = select_order_resp
        select_eq.order.return_value = select_order
        select_mock.eq.return_value = select_eq

        # ---- UPDATE branch -------------------------------------------------
        def update_side_effect(payload: dict[str, Any]) -> MagicMock:
            update_mock = MagicMock()
            update_eq = MagicMock()
            update_resp = MagicMock()
            update_resp.data = [{"id": "ignored"}]

            def eq_side_effect(col: str, val: Any) -> MagicMock:
                # Capture: (invoice_item_id → updated payload). The service
                # always filters by ``.eq("id", iid)`` before ``execute()``.
                captured.append({"id": val, "updates": payload})
                update_eq.execute.return_value = update_resp
                return update_eq

            update_mock.eq.side_effect = eq_side_effect
            return update_mock

        table_mock = MagicMock()
        table_mock.select.return_value = select_mock
        table_mock.update.side_effect = update_side_effect
        return table_mock

    mock_sb.table.side_effect = table_side_effect
    return mock_sb


def _make_invoice_item(
    item_id: str,
    idn_sku: str,
    *,
    brand: str = "SKF",
    quantity: int = 10,
    price: float | None = None,
) -> dict[str, Any]:
    """One invoice_items row with the coverage→quote_items embed populated."""
    return {
        "id": item_id,
        "invoice_id": "inv-001",
        "brand": brand,
        "supplier_sku": None,
        "product_name": "stub",
        "quantity": quantity,
        "minimum_order_quantity": None,
        "purchase_price_original": price,
        "production_time_days": None,
        "weight_in_kg": None,
        "dimension_height_mm": None,
        "dimension_width_mm": None,
        "dimension_length_mm": None,
        "supplier_notes": None,
        "invoice_item_coverage": [
            {
                "quote_item_id": f"qi-{item_id}",
                "ratio": 1,
                "quote_items": {"idn_sku": idn_sku},
            }
        ],
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestImportInvoiceXls:
    """Happy path + the 4 edge cases from the decisions doc."""

    @patch("services.xls_import_service.get_supabase")
    def test_happy_path_updates_all_matched_articles(self, mock_get_sb):
        """3 articles in XLS, 3 in КПП — all matched, all 3 updated."""
        captured: list[dict[str, Any]] = []
        mock_get_sb.return_value = _build_mock_supabase(
            invoice_items=[
                _make_invoice_item("ii-1", "SKU-A"),
                _make_invoice_item("ii-2", "SKU-B"),
                _make_invoice_item("ii-3", "SKU-C"),
            ],
            updates_capture=captured,
        )

        xlsx = make_xlsx(
            [
                {"idn_sku": "SKU-A", "purchase_price_original": 100, "quantity": 5},
                {"idn_sku": "SKU-B", "purchase_price_original": 200, "quantity": 8},
                {"idn_sku": "SKU-C", "purchase_price_original": 300, "quantity": 12},
            ]
        )

        from services.xls_import_service import import_invoice_xls

        result = import_invoice_xls(invoice_id="inv-001", file_bytes=xlsx)

        assert result["updated"] == 3
        assert result["skipped"] == []
        assert result["total_in_file"] == 3
        # Every invoice_item must have been touched
        assert {c["id"] for c in captured} == {"ii-1", "ii-2", "ii-3"}

    @patch("services.xls_import_service.get_supabase")
    def test_new_article_in_xls_not_in_kpp_is_skipped(self, mock_get_sb):
        """Article in XLS but not in КПП → skipped, never sent as UPDATE."""
        captured: list[dict[str, Any]] = []
        mock_get_sb.return_value = _build_mock_supabase(
            invoice_items=[_make_invoice_item("ii-1", "SKU-A")],
            updates_capture=captured,
        )

        xlsx = make_xlsx(
            [
                {"idn_sku": "SKU-A", "purchase_price_original": 100, "quantity": 5},
                {"idn_sku": "SKU-UNKNOWN", "purchase_price_original": 999, "quantity": 1},
            ]
        )

        from services.xls_import_service import import_invoice_xls

        result = import_invoice_xls(invoice_id="inv-001", file_bytes=xlsx)

        assert result["updated"] == 1
        assert result["skipped"] == ["SKU-UNKNOWN"]
        assert result["total_in_file"] == 2
        # The UNKNOWN row must not have triggered an UPDATE
        assert len(captured) == 1
        assert captured[0]["id"] == "ii-1"

    @patch("services.xls_import_service.get_supabase")
    def test_article_in_kpp_missing_from_xls_is_preserved(self, mock_get_sb):
        """Article in КПП but not in XLS → no UPDATE issued for that row."""
        captured: list[dict[str, Any]] = []
        mock_get_sb.return_value = _build_mock_supabase(
            invoice_items=[
                _make_invoice_item("ii-1", "SKU-A", price=50.0),
                _make_invoice_item("ii-2", "SKU-B", price=60.0),
            ],
            updates_capture=captured,
        )

        # Only SKU-A is in the file — SKU-B should be left untouched.
        xlsx = make_xlsx(
            [{"idn_sku": "SKU-A", "purchase_price_original": 100, "quantity": 5}]
        )

        from services.xls_import_service import import_invoice_xls

        result = import_invoice_xls(invoice_id="inv-001", file_bytes=xlsx)

        assert result["updated"] == 1
        assert result["skipped"] == []
        assert result["total_in_file"] == 1
        # Only ii-1 got updated — ii-2 was preserved.
        assert {c["id"] for c in captured} == {"ii-1"}

    @patch("services.xls_import_service.get_supabase")
    def test_duplicate_articles_in_xls_raises(self, mock_get_sb):
        """Same idn_sku twice in XLS → DuplicateArticlesError with the list."""
        captured: list[dict[str, Any]] = []
        mock_get_sb.return_value = _build_mock_supabase(
            invoice_items=[_make_invoice_item("ii-1", "SKU-A")],
            updates_capture=captured,
        )

        xlsx = make_xlsx(
            [
                {"idn_sku": "SKU-A", "purchase_price_original": 100, "quantity": 5},
                {"idn_sku": "SKU-A", "purchase_price_original": 200, "quantity": 8},
            ]
        )

        from services.xls_import_service import (
            DuplicateArticlesError,
            import_invoice_xls,
        )

        with pytest.raises(DuplicateArticlesError) as exc_info:
            import_invoice_xls(invoice_id="inv-001", file_bytes=xlsx)

        assert exc_info.value.duplicates == ["SKU-A"]
        # No updates were issued for the rejected file
        assert captured == []

    @patch("services.xls_import_service.get_supabase")
    def test_all_template_fields_get_updated(self, mock_get_sb):
        """Verify every column from the template lands in the UPDATE payload."""
        captured: list[dict[str, Any]] = []
        mock_get_sb.return_value = _build_mock_supabase(
            invoice_items=[_make_invoice_item("ii-1", "SKU-A")],
            updates_capture=captured,
        )

        xlsx = make_xlsx(
            [
                {
                    "brand": "FAG",
                    "idn_sku": "SKU-A",
                    "supplier_sku": "FAG-22220",
                    "product_name": "Подшипник роликовый",
                    "quantity": 7,
                    "minimum_order_quantity": 3,
                    "purchase_price_original": 250.5,
                    "production_time_days": 45,
                    "weight_in_kg": 3.2,
                    "dimensions": "10×20×30",
                    "supplier_notes": "Доставка через 2 недели",
                }
            ]
        )

        from services.xls_import_service import import_invoice_xls

        result = import_invoice_xls(invoice_id="inv-001", file_bytes=xlsx)

        assert result["updated"] == 1
        assert len(captured) == 1
        payload = captured[0]["updates"]
        assert payload["brand"] == "FAG"
        assert payload["supplier_sku"] == "FAG-22220"
        assert payload["product_name"] == "Подшипник роликовый"
        assert payload["quantity"] == 7
        assert payload["minimum_order_quantity"] == 3
        assert payload["purchase_price_original"] == 250.5
        assert payload["production_time_days"] == 45
        assert payload["weight_in_kg"] == 3.2
        # Dimensions parsed into 3 separate columns (mirrors handsontable save)
        assert payload["dimension_height_mm"] == 10
        assert payload["dimension_width_mm"] == 20
        assert payload["dimension_length_mm"] == 30
        assert payload["supplier_notes"] == "Доставка через 2 недели"

    @patch("services.xls_import_service.get_supabase")
    def test_missing_idn_sku_in_row_is_skipped(self, mock_get_sb):
        """Empty `Арт. запрошенный` cell → row is silently ignored.

        Defensive: users sometimes paste a half-empty row at the bottom of the
        sheet. We must not blow up on it, and we must not count it toward the
        skipped list (since there's no article to report back).
        """
        captured: list[dict[str, Any]] = []
        mock_get_sb.return_value = _build_mock_supabase(
            invoice_items=[_make_invoice_item("ii-1", "SKU-A")],
            updates_capture=captured,
        )

        xlsx = make_xlsx(
            [
                {"idn_sku": "SKU-A", "purchase_price_original": 100, "quantity": 5},
                # Empty article cell — should be dropped silently
                {"idn_sku": None, "purchase_price_original": 999, "quantity": 1},
            ]
        )

        from services.xls_import_service import import_invoice_xls

        result = import_invoice_xls(invoice_id="inv-001", file_bytes=xlsx)

        assert result["updated"] == 1
        # The blank row must not appear in skipped — only true mismatches do
        assert result["skipped"] == []
        # total_in_file counts only rows with a non-empty article
        assert result["total_in_file"] == 1
