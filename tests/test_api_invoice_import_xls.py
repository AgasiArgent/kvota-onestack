"""Tests for POST /api/invoices/{id}/import-xls (Testing 2 row 70).

Covers:
  - route registered on the FastAPI sub-app + reachable via the /api mount
  - 401 when no JWT
  - 400 with code=DUPLICATES when the file has duplicate articles
  - 200 with a structured summary on the happy path
"""

from __future__ import annotations

import os
import sys
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from starlette.testclient import TestClient

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from api.app import api_sub_app  # noqa: E402


HEADERS_RU = [
    "Бренд",
    "Арт. запрошенный",
    "Арт. производителя",
    "Наименование производителя",
    "Наименование",
    "Кол-во",
    "Ед. изм.",
    "Мин. заказ",
    "Цена",
    "Срок (к.д.)",
    "Вес (кг)",
    "Размеры (В×Ш×Д мм)",
    "Примечание",
    "Покрывает",
]


def _make_xlsx(rows: list[dict[str, object]]) -> bytes:
    """Build a minimal XLS file mirroring the export template column order.

    Testing 2 row 88: the article column carries ``product_code`` (the match
    key) and "Ед. изм." sits between "Кол-во" and "Мин. заказ", shifting
    every subsequent column by 1.
    """
    wb = Workbook()
    ws = wb.active
    for col_idx, header in enumerate(HEADERS_RU, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=2, value=row.get("product_code"))
        ws.cell(row=row_idx, column=6, value=row.get("quantity"))
        # Column 7 is "Ед. изм." (read-only on import); column 9 is "Цена".
        ws.cell(row=row_idx, column=9, value=row.get("price"))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def subapp_client() -> TestClient:
    return TestClient(api_sub_app)


class TestImportXlsRouteRegistration:
    """Route must be wired on the sub-app + appear in OpenAPI."""

    def test_route_registered(self, subapp_client: TestClient) -> None:
        invoice_id = "11111111-1111-1111-1111-111111111111"
        files = {"file": ("test.xlsx", b"dummy", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        response = subapp_client.post(
            f"/invoices/{invoice_id}/import-xls", files=files
        )
        # No JWT → 401. 404 would mean the route is missing.
        assert response.status_code != 404, (
            "POST /invoices/{id}/import-xls returned 404 — route not registered."
        )

    def test_appears_in_openapi(self, subapp_client: TestClient) -> None:
        response = subapp_client.get("/openapi.json")
        assert response.status_code == 200
        paths = response.json().get("paths", {})
        assert "/invoices/{invoice_id}/import-xls" in paths
        assert "post" in paths["/invoices/{invoice_id}/import-xls"]


class TestImportXlsAuth:
    """Auth gate: no JWT → 401, no procurement role → 403."""

    def test_returns_401_when_no_jwt(self, subapp_client: TestClient) -> None:
        invoice_id = "11111111-1111-1111-1111-111111111111"
        files = {
            "file": (
                "test.xlsx",
                b"dummy",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = subapp_client.post(
            f"/invoices/{invoice_id}/import-xls", files=files
        )
        assert response.status_code == 401


class TestImportXlsContract:
    """End-to-end contract test with mocked auth + supabase."""

    def _bypass_auth(self):
        """Replace _get_procurement_user with a happy-path stub."""
        return patch(
            "api.invoices._get_procurement_user",
            return_value=(
                {"id": "user-1", "org_id": "org-1", "role_slugs": {"procurement"}},
                None,
            ),
        )

    def _bypass_ownership(self, invoice_id: str):
        return patch(
            "api.invoices._verify_invoice_ownership",
            return_value=(
                {
                    "id": invoice_id,
                    "quote_id": "q-1",
                    "invoice_number": "INV-001",
                    "quotes": {"organization_id": "org-1"},
                },
                None,
            ),
        )

    def test_returns_400_on_duplicates(self, subapp_client: TestClient) -> None:
        """File with same article twice → 400 + DUPLICATES envelope."""
        invoice_id = "11111111-1111-1111-1111-111111111111"
        xlsx = _make_xlsx(
            [
                {"product_code": "ART-A", "quantity": 1, "price": 10},
                {"product_code": "ART-A", "quantity": 2, "price": 20},
            ]
        )

        # No DB lookup needed — duplicates are caught BEFORE the kpp fetch.
        # _build_mock_supabase returns an empty inventory; this is fine
        # because the early-return path doesn't touch supabase at all.
        with self._bypass_auth(), self._bypass_ownership(invoice_id), patch(
            "services.xls_import_service.get_supabase"
        ) as mock_sb:
            mock_sb.return_value = MagicMock()
            response = subapp_client.post(
                f"/invoices/{invoice_id}/import-xls",
                files={
                    "file": (
                        "test.xlsx",
                        xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

        assert response.status_code == 400, response.text
        body = response.json()
        assert body["success"] is False
        assert body["error"]["code"] == "DUPLICATES"
        # Duplicates list must surface in the error so the toast can render it
        assert "ART-A" in body["error"]["message"] or "ART-A" in str(body)

    def test_returns_200_on_happy_path(self, subapp_client: TestClient) -> None:
        """Happy path: matched article → updated count + structured summary."""
        invoice_id = "11111111-1111-1111-1111-111111111111"
        xlsx = _make_xlsx(
            [{"product_code": "ART-A", "quantity": 5, "price": 100}]
        )

        # Mock supabase to return one invoice_item that matches ART-A.
        def make_mock_sb():
            mock = MagicMock()

            def table_side_effect(name):
                table_mock = MagicMock()
                # SELECT — list current invoice_items
                select_eq_order_resp = MagicMock()
                select_eq_order_resp.data = [
                    {
                        "id": "ii-1",
                        "invoice_item_coverage": [
                            {
                                "quote_item_id": "qi-1",
                                "ratio": 1,
                                "quote_items": {"product_code": "ART-A"},
                            }
                        ],
                    }
                ]
                select_eq_order = MagicMock()
                select_eq_order.execute.return_value = select_eq_order_resp
                select_eq = MagicMock()
                select_eq.order.return_value = select_eq_order
                select_mock = MagicMock()
                select_mock.eq.return_value = select_eq
                table_mock.select.return_value = select_mock
                # UPDATE — capture nothing here; we only assert on the
                # endpoint's response envelope.
                update_mock = MagicMock()
                update_eq = MagicMock()
                update_resp = MagicMock()
                update_resp.data = [{"id": "ii-1"}]
                update_eq.execute.return_value = update_resp
                update_mock.eq.return_value = update_eq
                table_mock.update.return_value = update_mock
                return table_mock

            mock.table.side_effect = table_side_effect
            return mock

        with self._bypass_auth(), self._bypass_ownership(invoice_id), patch(
            "services.xls_import_service.get_supabase"
        ) as mock_sb:
            mock_sb.return_value = make_mock_sb()
            response = subapp_client.post(
                f"/invoices/{invoice_id}/import-xls",
                files={
                    "file": (
                        "test.xlsx",
                        xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

        assert response.status_code == 200, response.text
        body = response.json()
        assert body["success"] is True
        assert body["data"]["updated"] == 1
        assert body["data"]["skipped"] == []
        assert body["data"]["total_in_file"] == 1
