"""
Export Validation Service - Generates Excel with API vs Excel comparison

Creates an Excel file with:
1. API_Inputs tab - All uploaded input values
2. расчет tab - Modified to reference API_Inputs (formulas recalculate)
3. API_Results tab - API calculation outputs
4. Conditional formatting highlighting differences > 0.01%
"""

import io
import logging
import os
from copy import copy
from decimal import Decimal
from typing import Any, Dict, List, Optional

import openpyxl

from services.currency_service import convert_amount
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

logger = logging.getLogger(__name__)

# Template file path (relative to backend directory for Railway deployment)
# Railway deploys only the backend/ directory, so templates must be inside backend/
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "templates",
    "validation",
    "test_raschet_new_template_vat22.xlsm"
)

# Highlight style for differences > 0.01%
DIFF_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Thin border style
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# VALUE MAPPINGS - Convert API values to Excel-expected values
# =============================================================================

# Offer sale type mapping (API -> Russian dropdown values)
SALE_TYPE_MAP = {
    # English/code values
    "openbook": "поставка",
    "supply": "поставка",
    "transit": "транзит",
    "fin_transit": "финтранзит",
    "export": "экспорт",
    # Russian values (passthrough)
    "поставка": "поставка",
    "транзит": "транзит",
    "финтранзит": "финтранзит",
    "экспорт": "экспорт",
}

# Supplier country mapping (code -> Russian dropdown values)
# Includes both 2-letter codes and Russian names (for passthrough)
# IMPORTANT: Must map API values to Excel's list_vat table values for correct VLOOKUP
COUNTRY_MAP = {
    # 2-letter country codes
    "TR": "Турция",
    "CN": "Китай",
    "RU": "Россия",
    "DE": "Германия",
    "IT": "Италия",
    "LT": "Литва",
    "LV": "Латвия",
    "BG": "Болгария",
    "PL": "Польша",
    "AE": "ОАЭ",
    # Russian names (passthrough)
    "Турция": "Турция",
    "Китай": "Китай",
    "Россия": "Россия",
    "Германия": "Германия",
    "Италия": "Италия",
    "Литва": "Литва",
    "Латвия": "Латвия",
    "Болгария": "Болгария",
    "Польша": "Польша",
    "ОАЭ": "ОАЭ",
    "Прочие": "Прочие",
    # Excel passthrough (exact match with list_vat table)
    "ЕС (закупка между странами ЕС)": "ЕС (закупка между странами ЕС)",
    "Турция (отгрузка на транзитной зоне)": "Турция (отгрузка на транзитной зоне)",
    # API value → Excel value mappings (2026-01-28: Fix internal_markup VLOOKUP mismatch)
    # These map calculation_mapper.py normalized values to Excel's list_vat table values
    "Турция (транзитная зона)": "Турция (отгрузка на транзитной зоне)",
    "ЕС (между странами ЕС)": "ЕС (закупка между странами ЕС)",
}

# Fields that are percentages (need to be divided by 100)
PERCENTAGE_FIELDS = {
    "advance_to_supplier",
    "advance_from_client",
    "advance_on_loading",
    "advance_on_shipping",
    "advance_on_customs",
    "rate_forex_risk",
    "supplier_discount",
    "import_tariff",
    "markup",
}

# Fields that are monetary values (need USD → quote currency conversion)
# These are stored in USD internally but Excel expects quote currency
MONETARY_INPUT_FIELDS = {
    "logistics_supplier_hub",
    "logistics_hub_customs",
    "logistics_customs_client",
    "brokerage_hub",
    "brokerage_customs",
    "warehousing",
    "documentation",
    "other_costs",
}


# =============================================================================
# CELL MAPPINGS - Quote Level Inputs
# =============================================================================

QUOTE_INPUT_MAPPING = {
    # Cell in расчет -> (API field path, display name)
    "D5": ("seller_company", "Компания-продавец"),
    "D6": ("offer_sale_type", "Вид КП"),
    "D7": ("incoterms", "Базис поставки"),
    "D8": ("quote_currency", "Валюта КП"),
    "D9": ("delivery_time", "Срок поставки (дней)"),
    "D11": ("advance_to_supplier", "Аванс поставщику (%)"),
    # Payment milestones
    "J5": ("advance_from_client", "Аванс от клиента (%)"),
    "K5": ("time_to_advance", "Дней до аванса"),
    "J6": ("advance_on_loading", "Аванс при заборе (%)"),
    "K6": ("time_to_advance_loading", "Дней до аванса загрузки"),
    "J7": ("advance_on_shipping", "Аванс при отправке (%)"),
    "K7": ("time_to_advance_shipping", "Дней до аванса отправки"),
    "J8": ("advance_on_customs", "Аванс при таможне (%)"),
    "K8": ("time_to_advance_customs", "Дней до аванса таможни"),
    "K9": ("time_to_advance_on_receiving", "Дней до оплаты после получения"),
    # Logistics costs
    "W2": ("logistics_supplier_hub", "Логистика: Поставщик-Хаб"),
    "W3": ("logistics_hub_customs", "Логистика: Хаб-Таможня"),
    "W4": ("logistics_customs_client", "Логистика: Таможня-Клиент"),
    # Brokerage costs
    "W5": ("brokerage_hub", "Брокерские услуги: Хаб"),
    "W6": ("brokerage_customs", "Брокерские услуги: Таможня"),
    "W7": ("warehousing", "Расходы на СВХ"),
    "W8": ("documentation", "Разрешительные документы"),
    "W9": ("other_costs", "Прочее"),
    # DM Fee - type determines placement:
    # AG3: dm_fee_type ("Фикс" or "комиссия %")
    # AG4: dm_fee_value when type is "Фикс" (fixed amount)
    # AG6: dm_fee_value when type is "комиссия %" (percentage, divided by 100)
    "AG3": ("dm_fee_type", "Тип вознаграждения ЛПР"),
    # NOTE: dm_fee_value is handled separately in _modify_raschet_references()
    # Admin settings - C30/D30 on API_Inputs, mapped to расчет internally
    "D30": ("rate_forex_risk", "Резерв на курсовую разницу (%)"),
}

# =============================================================================
# CELL MAPPINGS - Product Level Inputs (Row 16 = first product)
# =============================================================================

PRODUCT_INPUT_COLUMNS = {
    # Column letter -> (API field, display name)
    "B": ("brand", "Бренд"),
    "C": ("sku", "Артикул"),
    "D": ("name", "Название товара"),
    "E": ("quantity", "Количество"),
    "G": ("weight_in_kg", "Вес, кг"),
    "J": ("currency_of_base_price", "Валюта закупки"),
    "K": ("base_price_vat", "Цена закупки (с VAT)"),
    "L": ("supplier_country", "Страна закупки"),
    "O": ("supplier_discount", "Скидка поставщика (%)"),
    "Q": ("exchange_rate", "Курс к валюте КП"),
    "W": ("customs_code", "Код ТН ВЭД"),
    "X": ("import_tariff", "Пошлина (%)"),
    "AC": ("markup", "Наценка (%)"),
}

# =============================================================================
# CELL MAPPINGS - Product Level Outputs (Calculated)
# =============================================================================

PRODUCT_OUTPUT_COLUMNS = {
    # Column letter -> (API field from ProductCalculationResult, display name)
    "N": ("purchase_price_no_vat", "Цена без VAT"),
    "P": ("purchase_price_after_discount", "После скидки"),
    "R": ("purchase_price_per_unit_quote_currency", "Цена за ед. в валюте КП"),
    "S": ("purchase_price_total_quote_currency", "Стоимость закупки"),
    "T": ("logistics_first_leg", "Логистика первый этап"),
    "U": ("logistics_last_leg", "Логистика второй этап"),
    "V": ("logistics_total", "Логистика итого"),
    "Y": ("customs_fee", "Пошлина (сумма)"),
    "Z": ("excise_tax_amount", "Акциз (сумма)"),
    "AA": ("cogs_per_unit", "Себестоимость за ед."),
    "AB": ("cogs_per_product", "Себестоимость товара"),
    "AD": ("sale_price_per_unit_excl_financial", "Цена продажи за ед. (без фин.)"),
    "AE": ("sale_price_total_excl_financial", "Цена продажи итого (без фин.)"),
    "AF": ("profit", "Прибыль"),
    "AG": ("dm_fee", "Вознаграждение ЛПР"),
    "AH": ("forex_reserve", "Резерв на курсовую разницу"),
    "AI": ("financial_agent_fee", "Комиссия ФинАгента"),
    "AJ": ("sales_price_per_unit_no_vat", "Цена продажи за ед. (без НДС)"),
    "AK": ("sales_price_total_no_vat", "Цена продажи итого (без НДС)"),
    "AL": ("sales_price_total_with_vat", "Цена продажи итого (с НДС)"),
    "AM": ("sales_price_per_unit_with_vat", "Цена продажи за ед. (с НДС)"),
    "AN": ("vat_from_sales", "НДС с продажи"),
    "AO": ("vat_on_import", "НДС к вычету"),
    "AP": ("vat_net_payable", "НДС к уплате"),
    "AQ": ("transit_commission", "Транзитная комиссия"),
    "AX": ("internal_sale_price_per_unit", "Внутренняя цена за ед."),
    "AY": ("internal_sale_price_total", "Внутренняя стоимость"),
    "BA": ("financing_cost_initial", "Начальное финансирование"),
    "BB": ("financing_cost_credit", "Проценты по отсрочке"),
}

# =============================================================================
# CELL MAPPINGS - Quote Level Totals (Row 13)
# =============================================================================

QUOTE_TOTAL_CELLS = {
    # Cell -> (API field, display name)
    "S13": ("total_purchase_price", "Итого стоимость закупки"),
    "T13": ("total_logistics_first", "Логистика первый этап (итого)"),
    "U13": ("total_logistics_last", "Логистика второй этап (итого)"),
    "V13": ("total_logistics", "Логистика итого"),
    "AB13": ("total_cogs", "Себестоимость итого"),
    "AK13": ("total_revenue", "Выручка (без НДС)"),
    "AL13": ("total_revenue_with_vat", "Выручка (с НДС)"),
    "AF13": ("total_profit", "Прибыль итого"),
}

# =============================================================================
# CELL MAPPINGS - Financing Outputs
# =============================================================================

FINANCING_CELLS = {
    # Cell -> (API field, display name)
    "BH2": ("evaluated_revenue", "Оценочная выручка"),
    "BH3": ("client_advance", "Аванс клиента"),
    "BH4": ("total_before_forwarding", "Итого до экспедирования"),
    "BH6": ("supplier_payment", "Платеж поставщику"),
    "BJ7": ("supplier_financing_cost", "Стоимость фин-ия поставщика"),
    "BJ10": ("operational_financing_cost", "Стоимость операционного фин-ия"),
    "BJ11": ("total_financing_cost", "Итого стоимость финансирования"),
    "BL3": ("credit_sales_amount", "Сумма к оплате клиентом"),
    "BL4": ("credit_sales_fv", "FV с процентами"),
    "BL5": ("credit_sales_interest", "Проценты по отсрочке"),
}


class ExportValidationService:
    """Service to generate validation Excel with API vs Excel comparison."""

    def __init__(self, template_path: str = TEMPLATE_PATH):
        self.template_path = template_path

    def generate_validation_export(
        self,
        quote_inputs: Dict[str, Any],
        product_inputs: List[Dict[str, Any]],
        api_results: Dict[str, Any],
        product_results: List[Dict[str, Any]],
    ) -> bytes:
        """
        Generate validation Excel file.

        Args:
            quote_inputs: Quote-level input values
            product_inputs: List of product input dicts
            api_results: Quote-level API results (totals, financing)
            product_results: List of product result dicts from API

        Returns:
            Excel file as bytes
        """
        # Load template
        wb = openpyxl.load_workbook(self.template_path, keep_vba=True)

        # 1. Create API_Inputs sheet
        self._create_inputs_sheet(wb, quote_inputs, product_inputs)

        # 2. Modify расчет to reference API_Inputs
        self._modify_raschet_references(wb, len(product_inputs))

        # 3. Create API_Results sheet
        self._create_results_sheet(wb, api_results, product_results)

        # 4. Add conditional formatting for comparison
        self._add_comparison_formatting(wb, len(product_results))

        # 5. Protect all sheets with password
        self._protect_all_sheets(wb, password="vba2025")

        # 6. Force formula recalculation on open
        # This ensures date-based formulas (like BH2 VAT multiplier) recalculate
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_inputs_sheet(
        self,
        wb: openpyxl.Workbook,
        quote_inputs: Dict[str, Any],
        product_inputs: List[Dict[str, Any]],
    ) -> None:
        """Create API_Inputs sheet with all uploaded values.

        Monetary values (logistics, brokerage) are stored in USD internally.
        This sheet adds conversion formulas to convert them to quote currency
        for the расчет sheet to reference.
        """

        # Create or get sheet
        if "API_Inputs" in wb.sheetnames:
            ws = wb["API_Inputs"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("API_Inputs", 0)

        # Header
        ws["A1"] = "API Input Values"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Exchange rate info (row 2) - used by formulas below
        # Get rate from quote_inputs (passed from quotes_upload.py)
        usd_to_quote_rate = quote_inputs.get("_usd_to_quote_rate", 1.0)
        quote_currency = quote_inputs.get("_quote_currency", "USD")

        ws["A2"] = "Exchange Rate:"
        ws["B2"] = f"1 USD = {usd_to_quote_rate:.4f} {quote_currency}"
        ws["B2"].font = Font(bold=True, color="0000FF")
        ws["E2"] = usd_to_quote_rate  # Rate cell for formulas (E2)
        ws["E2"].number_format = '0.0000'
        ws["F2"] = "← Rate cell (E2)"
        ws["F2"].font = Font(color="888888", italic=True)

        # Store for later use
        self._usd_to_quote_rate = usd_to_quote_rate
        self._quote_currency = quote_currency

        # Section: Quote Settings
        row = 4
        ws[f"A{row}"] = "Quote Settings"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        ws[f"B{row}"].fill = HEADER_FILL
        ws[f"C{row}"].fill = HEADER_FILL
        ws[f"D{row}"].fill = HEADER_FILL
        row += 1

        ws[f"A{row}"] = "Cell"
        ws[f"B{row}"] = "Field"
        ws[f"C{row}"] = "Value (USD)"
        ws[f"D{row}"] = f"Value ({quote_currency})"
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].border = THIN_BORDER
        row += 1

        # Quote input values
        quote_input_start_row = row
        for cell_addr, (field, display_name) in QUOTE_INPUT_MAPPING.items():
            ws[f"A{row}"] = cell_addr
            ws[f"B{row}"] = display_name
            value = quote_inputs.get(field)
            # Use field-aware formatting (percentages, enums)
            ws[f"C{row}"] = self._format_input_value(value, field)
            ws[f"C{row}"].alignment = Alignment(horizontal="right")

            # Column D: For monetary fields, add conversion formula
            # For non-monetary fields, just reference C (passthrough)
            if field in MONETARY_INPUT_FIELDS:
                # Formula: =C{row}*$E$2 (USD value * exchange rate)
                ws[f"D{row}"] = f"=C{row}*$E$2"
                ws[f"D{row}"].number_format = '#,##0.00'
            else:
                # Non-monetary: passthrough (reference C directly)
                ws[f"D{row}"] = f"=C{row}"

            ws[f"D{row}"].alignment = Alignment(horizontal="right")
            for col in ["A", "B", "C", "D"]:
                ws[f"{col}{row}"].border = THIN_BORDER
            row += 1

        # Store advance_from_client for D10 calculation
        self._advance_from_client = quote_inputs.get("advance_from_client", 0)

        # Handle dm_fee_value separately - placement depends on type
        dm_fee_type = quote_inputs.get("dm_fee_type", "")
        dm_fee_value = quote_inputs.get("dm_fee_value", 0)

        # Determine target cell based on type:
        # "Фикс" or "фикс. сумма" -> AG4 (fixed amount)
        # "комиссия %" or "% от суммы" -> AG6 (percentage, divide by 100)
        is_percentage = any(pct in str(dm_fee_type).lower() for pct in ["%", "комиссия", "процент"])

        if is_percentage:
            dm_target_cell = "AG6"
            # Treat as percentage - divide by 100
            formatted_value = dm_fee_value / 100 if dm_fee_value else 0
        else:
            dm_target_cell = "AG4"
            formatted_value = dm_fee_value

        # Add dm_fee_value to API_Inputs
        ws[f"A{row}"] = dm_target_cell
        ws[f"B{row}"] = "Вознаграждение ЛПР"
        ws[f"C{row}"] = formatted_value
        ws[f"C{row}"].alignment = Alignment(horizontal="right")
        # For fixed DM fee: convert from USD to quote currency (=C*$E$2)
        # For percentage: passthrough (no conversion needed)
        if is_percentage:
            ws[f"D{row}"] = f"=C{row}"  # Percentage - passthrough
        else:
            ws[f"D{row}"] = f"=C{row}*$E$2"  # Fixed amount - convert USD to quote currency
            ws[f"D{row}"].number_format = '#,##0.00'
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = THIN_BORDER

        # Store for _modify_raschet_references
        self._dm_fee_row = row
        self._dm_fee_target_cell = dm_target_cell
        row += 1

        # Section: Products
        row += 2
        product_header_row = row
        ws[f"A{row}"] = "Products"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        # Product headers
        product_start_row = row + 1
        ws[f"A{row}"] = "#"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = THIN_BORDER

        col_idx = 2
        product_col_map = {}  # column letter -> field name
        for col_letter, (field, display_name) in PRODUCT_INPUT_COLUMNS.items():
            cell = ws.cell(row=row, column=col_idx)
            cell.value = display_name
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            product_col_map[col_idx] = field
            col_idx += 1
        row += 1

        # Product data
        for i, product in enumerate(product_inputs, 1):
            ws[f"A{row}"] = i
            ws[f"A{row}"].border = THIN_BORDER

            col_idx = 2
            for _, (field, _) in PRODUCT_INPUT_COLUMNS.items():
                cell = ws.cell(row=row, column=col_idx)
                value = product.get(field)
                # Use field-aware formatting (percentages, country codes)
                cell.value = self._format_input_value(value, field)
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
                col_idx += 1
            row += 1

        # Store row references as instance attributes for later use
        self._quote_input_start = quote_input_start_row
        self._product_start_row = product_start_row

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        for col in range(5, col_idx + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _modify_raschet_references(
        self, wb: openpyxl.Workbook, num_products: int
    ) -> None:
        """Modify расчет sheet to reference API_Inputs.

        References column D (converted to quote currency) for all values.
        Column D contains:
        - For monetary fields: =C*$E$2 (USD * exchange rate)
        - For non-monetary fields: =C (passthrough)
        """

        ws = wb["расчет"]
        inputs_sheet = "API_Inputs"

        # Quote-level inputs - create reference formulas
        # NOTE: Row numbers adjusted for new header layout (row 4 = header, row 5 = column names, row 6+ = data)
        row = 6  # Start row in API_Inputs for quote values (after header rows)
        for cell_addr, (field, _) in QUOTE_INPUT_MAPPING.items():
            # Create formula referencing API_Inputs column D (converted values)
            ws[cell_addr] = f"='{inputs_sheet}'!D{row}"
            row += 1

        # D10 - Payment type based on advance_from_client
        # Keep as value, not formula reference (per user request)
        payment_type = self._get_payment_type_value(self._advance_from_client)
        ws["D10"] = payment_type

        # DM Fee value - placed in AG4 (Фикс) or AG6 (комиссия %)
        # The target cell and row were determined in _create_inputs_sheet
        # Reference column D (converted value)
        if hasattr(self, '_dm_fee_target_cell') and hasattr(self, '_dm_fee_row'):
            ws[self._dm_fee_target_cell] = f"='{inputs_sheet}'!D{self._dm_fee_row}"

        # Product-level inputs
        # Use the stored row from _create_inputs_sheet (row after headers)
        # Products start at row 16 in расчет
        inputs_product_start = self._product_start_row  # Set by _create_inputs_sheet
        raschet_product_start = 16

        for prod_idx in range(num_products):
            inputs_row = inputs_product_start + prod_idx
            raschet_row = raschet_product_start + prod_idx

            # Map each product input column
            input_col = 2  # Start at column B in API_Inputs
            for col_letter, (field, _) in PRODUCT_INPUT_COLUMNS.items():
                raschet_cell = f"{col_letter}{raschet_row}"
                inputs_cell = f"'{inputs_sheet}'!{get_column_letter(input_col)}{inputs_row}"
                ws[raschet_cell] = f"={inputs_cell}"
                input_col += 1

    def _create_results_sheet(
        self,
        wb: openpyxl.Workbook,
        api_results: Dict[str, Any],
        product_results: List[Dict[str, Any]],
    ) -> None:
        """Create API_Results sheet with calculation outputs.

        API results are in USD. Excel formulas convert to quote currency for comparison.
        """

        # Create or get sheet
        if "API_Results" in wb.sheetnames:
            ws = wb["API_Results"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("API_Results")

        # Get currency info first (needed for header)
        usd_to_quote_rate = api_results.get("_usd_to_quote_rate", 1.0)
        quote_currency = api_results.get("_quote_currency", "USD")

        # Header
        ws["A1"] = f"API Calculation Results ({quote_currency})"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        # Exchange rate info (row 2) - used by formulas below
        # Store rate in G2 for formula reference

        ws["A2"] = "Exchange Rate:"
        ws["B2"] = f"1 USD = {usd_to_quote_rate:.4f} {quote_currency}"
        ws["B2"].font = Font(bold=True, color="0000FF")
        ws["G2"] = usd_to_quote_rate  # Rate cell for formulas
        ws["G2"].number_format = '0.0000'
        ws["H2"] = "← Rate cell (G2)"
        ws["H2"].font = Font(color="888888", italic=True)

        # Section: Quote Totals
        row = 4
        ws[f"A{row}"] = "Quote Totals"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        ws[f"A{row}"] = "Cell"
        ws[f"B{row}"] = "Field"
        ws[f"C{row}"] = f"API ({quote_currency})"
        ws[f"D{row}"] = "For comparison"
        ws[f"E{row}"] = f"Excel ({quote_currency})"
        ws[f"F{row}"] = "Diff %"
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].border = THIN_BORDER
        row += 1

        # Quote totals - Most API values are in QUOTE CURRENCY
        # Only LOGISTICS totals are in USD (from invoice aggregation)
        #
        # Fields that need USD → quote currency conversion:
        QUOTE_TOTAL_USD_FIELDS = {
            "total_logistics_first",   # T13 - sum of T16:Txx (USD from invoices)
            "total_logistics_last",    # U13 - sum of U16:Uxx (USD from invoices)
            "total_logistics",         # V13 - sum of V16:Vxx (USD from invoices)
        }
        # Fields already in quote currency (no conversion needed)
        QUOTE_TOTAL_QUOTE_CURRENCY = {
            "total_purchase_price",    # S13 - already in quote currency
            "total_cogs",              # AB13 - already in quote currency
            "total_revenue",           # AK13 - already in quote currency
            "total_revenue_with_vat",  # AL13 - already in quote currency
            "total_profit",            # AF13 - already in quote currency
        }

        for cell_addr, (field, display_name) in QUOTE_TOTAL_CELLS.items():
            ws[f"A{row}"] = cell_addr
            ws[f"B{row}"] = display_name
            api_value = api_results.get(field)
            ws[f"C{row}"] = self._format_value(api_value)
            # D: Convert from USD to quote currency only for logistics
            if field in QUOTE_TOTAL_USD_FIELDS:
                ws[f"D{row}"] = f"=C{row}*API_Inputs!$E$2"
            else:
                ws[f"D{row}"] = f"=C{row}"  # Already in quote currency
            ws[f"D{row}"].number_format = '#,##0.00'
            ws[f"E{row}"] = f"=расчет!{cell_addr}"
            # Diff formula: compare D (API converted to quote currency) vs E (Excel in quote currency)
            ws[f"F{row}"] = f'=IF(E{row}=0,"N/A",ABS(D{row}-E{row})/ABS(E{row}))'
            ws[f"F{row}"].number_format = '0.00%'
            for col in ["A", "B", "C", "D", "E", "F"]:
                ws[f"{col}{row}"].border = THIN_BORDER
            row += 1

        # Financing cells - API values are in QUOTE CURRENCY (same as other outputs)
        row += 1
        ws[f"A{row}"] = "Financing"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        # Financing fields are already in quote currency (no conversion needed)
        # Engine calculates everything in quote currency except logistics
        FINANCING_QUOTE_CURRENCY_FIELDS = {
            "evaluated_revenue",           # BH2 - in quote currency
            "client_advance",              # BH3 - in quote currency
            "total_before_forwarding",     # BH4 - in quote currency
            "supplier_payment",            # BH6 - in quote currency
            "supplier_financing_cost",     # BJ7 - in quote currency
            "operational_financing_cost",  # BJ10 - in quote currency
            "total_financing_cost",        # BJ11 - in quote currency
            "credit_sales_amount",         # BL3 - in quote currency
            "credit_sales_fv",             # BL4 - in quote currency
            "credit_sales_interest",       # BL5 - in quote currency
        }

        for cell_addr, (field, display_name) in FINANCING_CELLS.items():
            ws[f"A{row}"] = cell_addr
            ws[f"B{row}"] = display_name
            api_value = api_results.get(field)
            ws[f"C{row}"] = self._format_value(api_value)
            # D: No conversion needed - already in quote currency
            ws[f"D{row}"] = f"=C{row}"
            ws[f"D{row}"].number_format = '#,##0.00'
            ws[f"E{row}"] = f"=расчет!{cell_addr}"
            ws[f"F{row}"] = f'=IF(E{row}=0,"N/A",ABS(D{row}-E{row})/ABS(E{row}))'
            ws[f"F{row}"].number_format = '0.00%'
            for col in ["A", "B", "C", "D", "E", "F"]:
                ws[f"{col}{row}"].border = THIN_BORDER
            row += 1

        # Section: Product Results
        # NOTE: Most values from calculation engine are in QUOTE CURRENCY
        # Only logistics (T, U, V) are in USD and converted in Comparison sheet
        row += 2
        ws[f"A{row}"] = f"Product Results (Quote Currency, logistics in USD)"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = HEADER_FILL
        row += 1

        # Product headers
        ws[f"A{row}"] = "#"
        ws[f"A{row}"].font = Font(bold=True)

        col_idx = 2
        for col_letter, (field, display_name) in PRODUCT_OUTPUT_COLUMNS.items():
            cell = ws.cell(row=row, column=col_idx)
            cell.value = f"{col_letter}: {display_name[:15]}"
            cell.font = Font(bold=True, size=9)
            cell.border = THIN_BORDER
            col_idx += 1
        row += 1

        product_results_start = row

        # Product data
        for i, product in enumerate(product_results):
            raschet_row = 16 + i  # Row in расчет sheet

            ws[f"A{row}"] = i + 1
            ws[f"A{row}"].border = THIN_BORDER

            col_idx = 2
            for col_letter, (field, _) in PRODUCT_OUTPUT_COLUMNS.items():
                cell = ws.cell(row=row, column=col_idx)
                api_value = product.get(field)
                cell.value = self._format_value(api_value)
                cell.alignment = Alignment(horizontal="right")
                cell.border = THIN_BORDER
                col_idx += 1
            row += 1

        # Store for conditional formatting
        self._product_results_start = product_results_start
        self._num_products = len(product_results)

        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10

    def _add_comparison_formatting(
        self, wb: openpyxl.Workbook, num_products: int
    ) -> None:
        """Add conditional formatting to highlight differences > 0.01%."""

        ws = wb["API_Results"]

        # Highlight F column (Diff %) where value > 0.01%
        # Sheet structure: C=API(USD), D=API(Quote), E=Excel(Quote), F=Diff%
        diff_rule = FormulaRule(
            formula=["F5>0.0001"],  # 0.01% = 0.0001 (F column contains decimal, not percent)
            fill=DIFF_FILL,
        )

        # Apply to diff column range
        ws.conditional_formatting.add("F5:F100", diff_rule)

        # Also add comparison sheet for detailed product analysis
        self._create_comparison_sheet(wb, num_products)

    def _create_comparison_sheet(
        self, wb: openpyxl.Workbook, num_products: int
    ) -> None:
        """Create detailed comparison sheet for products.

        API results are in USD, Excel (расчет) values are in quote currency.
        We convert API values to quote currency using the rate from API_Results!G2.
        """

        if "Comparison" in wb.sheetnames:
            ws = wb["Comparison"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("Comparison")

        ws["A1"] = "Detailed Comparison (Cells with >0.01% difference highlighted)"
        ws["A1"].font = Font(bold=True, size=12)
        ws.merge_cells("A1:G1")

        # Exchange rate reference note
        ws["A2"] = "Rate from API_Results!G2"
        ws["A2"].font = Font(color="888888", italic=True)

        row = 4
        quote_currency = self._quote_currency if hasattr(self, '_quote_currency') else "Quote"
        ws[f"A{row}"] = "Product"
        ws[f"B{row}"] = "Cell"
        ws[f"C{row}"] = "Field"
        ws[f"D{row}"] = f"API ({quote_currency})"
        ws[f"E{row}"] = "For comparison"
        ws[f"F{row}"] = f"Excel ({quote_currency})"
        ws[f"G{row}"] = "Diff %"
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].fill = HEADER_FILL
            ws[f"{col}{row}"].font = HEADER_FONT
            ws[f"{col}{row}"].border = THIN_BORDER
        row += 1

        # Currency handling in Comparison sheet:
        #
        # IMPORTANT: Calculation engine works in QUOTE CURRENCY internally.
        # Only LOGISTICS values come from invoice aggregation in USD.
        #
        # Categories:
        # 1. BASE currency fields (N16, P16) - supplier's currency, no conversion needed
        # 2. USD fields (ONLY logistics) - need USD → quote currency conversion
        # 3. Quote currency fields (most outputs) - already in quote currency, NO conversion
        # 4. Percentage/count fields - no conversion needed
        #
        # Conversion formula: =D{row}*API_Inputs!$E$2 (where E2 = USD to quote rate)
        #
        BASE_CURRENCY_FIELDS = {"purchase_price_no_vat", "purchase_price_after_discount"}

        # Fields that are in USD and need conversion to quote currency
        # Logistics comes from invoice aggregation in USD
        # DM Fee is passed to engine in USD
        USD_MONETARY_FIELDS = {
            "logistics_first_leg",                     # T16 - from invoices, USD
            "logistics_last_leg",                      # U16 - from invoices, USD
            "logistics_total",                         # V16 - from invoices, USD
            "dm_fee",                                  # AG16 - passed to engine in USD
        }

        # Fields already in quote currency (engine outputs in quote currency)
        # These should NOT be converted - passthrough only
        QUOTE_CURRENCY_FIELDS = {
            "purchase_price_per_unit_quote_currency",  # R16 = P16 / exchange_rate
            "purchase_price_total_quote_currency",     # S16 = R16 * quantity
            "customs_fee",                             # Y16 - calculated in quote currency
            "excise_tax_amount",                       # Z16 - calculated in quote currency
            "cogs_per_unit",                           # AA16 - sum of quote currency values
            "cogs_per_product",                        # AB16 - sum of quote currency values
            "sale_price_per_unit_excl_financial",      # AD16 - calculated in quote currency
            "sale_price_total_excl_financial",         # AE16 - calculated in quote currency
            "profit",                                  # AF16 - calculated in quote currency
            # dm_fee is in USD_MONETARY_FIELDS (passed to engine in USD)
            "forex_reserve",                           # AH16 - calculated in quote currency
            "financial_agent_fee",                     # AI16 - calculated in quote currency
            "sales_price_per_unit_no_vat",             # AJ16 - calculated in quote currency
            "sales_price_total_no_vat",                # AK16 - calculated in quote currency
            "sales_price_total_with_vat",              # AL16 - calculated in quote currency
            "sales_price_per_unit_with_vat",           # AM16 - calculated in quote currency
            "vat_from_sales",                          # AN16 - calculated in quote currency
            "vat_on_import",                           # AO16 - calculated in quote currency
            "vat_net_payable",                         # AP16 - calculated in quote currency
            "transit_commission",                      # AQ16 - calculated in quote currency
            "internal_sale_price_per_unit",            # AX16 - calculated in quote currency
            "internal_sale_price_total",               # AY16 - calculated in quote currency
            "financing_cost_initial",                  # BA16 - calculated in quote currency
            "financing_cost_credit",                   # BB16 - calculated in quote currency
        }

        # Backwards compatibility alias
        ALREADY_CONVERTED_FIELDS = QUOTE_CURRENCY_FIELDS

        # For each product and each output column
        for prod_idx in range(num_products):
            raschet_row = 16 + prod_idx
            api_results_row = self._product_results_start + prod_idx

            for col_letter, (field, display_name) in PRODUCT_OUTPUT_COLUMNS.items():
                ws[f"A{row}"] = prod_idx + 1
                ws[f"B{row}"] = f"{col_letter}{raschet_row}"
                ws[f"C{row}"] = display_name

                # Column D: API value from API_Results (in USD)
                api_col = list(PRODUCT_OUTPUT_COLUMNS.keys()).index(col_letter) + 2
                ws[f"D{row}"] = f"=API_Results!{get_column_letter(api_col)}{api_results_row}"

                # Column E: API value converted to quote currency for comparison
                # - BASE currency fields: passthrough (same currency for API and Excel)
                # - Already converted fields (R16, S16): passthrough
                # - USD monetary fields: multiply by USD→quote rate from API_Inputs!E2
                if field in BASE_CURRENCY_FIELDS or field in ALREADY_CONVERTED_FIELDS:
                    ws[f"E{row}"] = f"=D{row}"
                elif field in USD_MONETARY_FIELDS:
                    # Convert from USD to quote currency: value * rate
                    ws[f"E{row}"] = f"=D{row}*API_Inputs!$E$2"
                else:
                    # Unknown field - passthrough (might be percentage or count)
                    ws[f"E{row}"] = f"=D{row}"
                ws[f"E{row}"].number_format = '#,##0.00'

                # Column F: Excel value from расчет
                ws[f"F{row}"] = f"=расчет!{col_letter}{raschet_row}"

                # Column G: Diff percentage - compare E (API, possibly converted) vs F (Excel)
                ws[f"G{row}"] = f'=IF(F{row}=0,"N/A",ABS(E{row}-F{row})/ABS(F{row}))'
                ws[f"G{row}"].number_format = '0.00%'

                for col in ["A", "B", "C", "D", "E", "F", "G"]:
                    ws[f"{col}{row}"].border = THIN_BORDER
                row += 1

        # Conditional formatting for diff > 0.01%
        # G column contains diff as decimal (0.01% = 0.0001)
        diff_rule = FormulaRule(
            formula=["$G5>0.0001"],  # G column, starting at row 5 (after headers)
            fill=DIFF_FILL,
        )
        ws.conditional_formatting.add(f"A5:G{row}", diff_rule)

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 10

    def _format_value(self, value: Any) -> Any:
        """Format value for Excel cell."""
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, str):
            # Try to convert numeric strings
            try:
                return float(value)
            except ValueError:
                return value
        return value

    def _format_input_value(self, value: Any, field_name: str) -> Any:
        """
        Format input value for Excel, with field-aware transformations.

        - Percentages: 30 -> 0.3 (Excel expects decimal format)
        - Sale type: openbook -> поставка
        - Country: TR -> Турция
        """
        if value is None:
            return ""

        # Convert sale type to Russian
        if field_name == "offer_sale_type":
            str_val = str(value).lower()
            return SALE_TYPE_MAP.get(str_val, str(value))

        # Convert country code to Russian
        # Handle both 2-letter codes (TR) and Russian names (Турция)
        if field_name == "supplier_country":
            str_val = str(value)
            # First try exact match (handles Russian names)
            if str_val in COUNTRY_MAP:
                return COUNTRY_MAP[str_val]
            # Then try uppercase (handles 2-letter codes like "tr" -> "TR")
            upper_val = str_val.upper()
            if upper_val in COUNTRY_MAP:
                return COUNTRY_MAP[upper_val]
            # Return as-is if no match
            return str_val

        # Convert percentage fields (30 -> 0.3)
        if field_name in PERCENTAGE_FIELDS:
            try:
                num_val = float(value) if not isinstance(value, (int, float, Decimal)) else float(value)
                # If value > 1, assume it's in percent form (30 means 30%)
                if num_val > 1:
                    return num_val / 100
                return num_val
            except (ValueError, TypeError):
                return value

        # Default formatting
        return self._format_value(value)

    def _protect_all_sheets(self, wb: openpyxl.Workbook, password: str = "vba2025") -> None:
        """
        Protect all sheets in the workbook with the specified password,
        except for "КП open book" and "КП price" which remain unprotected.

        Args:
            wb: Workbook to protect
            password: Protection password (default: vba2025)
        """
        # Sheets to leave unprotected (for user editing)
        unprotected_sheets = {"КП open book", "КП price"}

        for sheet_name in wb.sheetnames:
            if sheet_name in unprotected_sheets:
                logger.debug(f"Skipping protection for sheet '{sheet_name}'")
                continue

            ws = wb[sheet_name]
            # Enable sheet protection
            ws.protection.sheet = True
            ws.protection.password = password
            # Standard protection settings - allow users to view but not edit
            ws.protection.enable()
            logger.debug(f"Protected sheet '{sheet_name}' with password")

    def _get_payment_type_value(self, advance_from_client: float) -> str:
        """
        Get payment type for D10 based on advance_from_client percentage.

        Returns:
            - "100% предоплата" if advance is 100% (or 1.0)
            - "100% постоплата" if advance is 0%
            - "частичная оплата" otherwise
        """
        # Normalize: if > 1, assume percent form
        if advance_from_client > 1:
            advance_from_client = advance_from_client / 100

        if advance_from_client >= 0.99:  # ~100%
            return "100% предоплата"
        elif advance_from_client <= 0.01:  # ~0%
            return "100% постоплата"
        else:
            return "частичная оплата"


# Convenience function
def generate_validation_export(
    quote_inputs: Dict[str, Any],
    product_inputs: List[Dict[str, Any]],
    api_results: Dict[str, Any],
    product_results: List[Dict[str, Any]],
) -> bytes:
    """Generate validation Excel file."""
    service = ExportValidationService()
    return service.generate_validation_export(
        quote_inputs, product_inputs, api_results, product_results
    )


# =============================================================================
# OneStack ADAPTER - Convert ExportData to validation service format
# =============================================================================

# Mapping for DM fee type (English -> Russian for Excel dropdown)
DM_FEE_TYPE_MAP = {
    # English values
    "fixed": "Фикс",
    "fix": "Фикс",
    "фикс": "Фикс",
    "фикс. сумма": "Фикс",
    # Percentage values
    "percent": "комиссия %",
    "percentage": "комиссия %",
    "%": "комиссия %",
    "комиссия": "комиссия %",
    "комиссия %": "комиссия %",
    "% от суммы": "комиссия %",
}

# Mapping for sale type (English -> Russian for Excel dropdown)
SALE_TYPE_ADAPTER_MAP = {
    "openbook": "поставка",
    "supply": "поставка",
    "transit": "транзит",
    "fin_transit": "финтранзит",
    "export": "экспорт",
    # Russian passthrough
    "поставка": "поставка",
    "транзит": "транзит",
    "финтранзит": "финтранзит",
    "экспорт": "экспорт",
}


def _map_dm_fee_type(value: str) -> str:
    """Map DM fee type to Russian value expected by Excel."""
    if not value:
        return "Фикс"
    return DM_FEE_TYPE_MAP.get(str(value).lower(), "Фикс")


def _map_sale_type(value: str) -> str:
    """Map sale type to Russian value expected by Excel."""
    if not value:
        return "поставка"
    return SALE_TYPE_ADAPTER_MAP.get(str(value).lower(), str(value))


def _get_usd_to_quote_rate(quote_currency: str) -> float:
    """
    Get exchange rate from USD to quote currency.

    For validation export, we need to convert API results (in USD) to quote currency.
    Returns: how many units of quote_currency = 1 USD

    Example: if quote is EUR and 1 USD = 0.92 EUR, returns 0.92
    """
    if not quote_currency or quote_currency == "USD":
        return 1.0

    try:
        # convert_amount(1 USD, from=USD, to=quote_currency)
        rate = convert_amount(Decimal("1"), "USD", quote_currency)
        return round(float(rate), 4) if rate else 1.0  # 4 decimal places per CBR standard
    except Exception as e:
        logger.warning(f"Could not get USD to {quote_currency} rate: {e}")
        return 1.0


def _get_exchange_rate_to_quote(from_currency: str, quote_currency: str) -> float:
    """
    Get exchange rate for Excel column Q (product-level).

    Excel formula R16 = P16 / Q16 divides purchase price by exchange rate.
    So Q must contain: "how many units of from_currency = 1 quote_currency"

    Example: if from_currency=USD, quote_currency=EUR:
    - Direct rate: 1 USD = 0.8553 EUR
    - Excel expects: 1 EUR = 1.169 USD (inverse)
    - Formula: 1000 USD / 1.169 = 855.26 EUR ✓

    Returns rate with 4 decimal places precision (CBR standard).
    """
    if not from_currency or not quote_currency:
        return 1.0
    if from_currency == quote_currency:
        return 1.0

    try:
        # Get direct rate: 1 from_currency = X quote_currency
        direct_rate = convert_amount(Decimal("1"), from_currency, quote_currency)
        if not direct_rate or float(direct_rate) == 0:
            return 1.0
        # Excel needs inverse: 1 quote_currency = Y from_currency
        inverse_rate = 1.0 / float(direct_rate)
        return round(inverse_rate, 4)
    except Exception as e:
        logger.warning(f"Could not get {from_currency} to {quote_currency} rate: {e}")
        return 1.0


def create_validation_excel(data) -> bytes:
    """
    Create validation Excel from OneStack ExportData.

    This adapter converts ExportData (from export_data_mapper.py) into
    the format expected by ExportValidationService.

    Args:
        data: ExportData from fetch_export_data()

    Returns:
        XLSM Excel file as bytes with macros
    """
    quote = data.quote
    items = data.items
    variables = data.variables
    calculations = data.calculations

    # Helper to convert value from source_currency to USD
    # Brokerage and DM Fee are stored in ORIGINAL currency, need conversion to USD for export
    def to_usd(value, currency):
        if not value or currency == 'USD':
            return value or 0
        try:
            val = Decimal(str(value))
            if val > 0:
                return float(convert_amount(val, currency, 'USD'))
        except:
            pass
        return value or 0

    # Convert brokerage fields from their original currencies to USD
    brokerage_hub_usd = to_usd(
        variables.get('brokerage_hub', 0),
        variables.get('brokerage_hub_currency', 'USD')
    )
    brokerage_customs_usd = to_usd(
        variables.get('brokerage_customs', 0),
        variables.get('brokerage_customs_currency', 'USD')
    )
    warehousing_usd = to_usd(
        variables.get('warehousing_at_customs', 0),
        variables.get('warehousing_at_customs_currency', 'USD')
    )
    # Debug: log raw values before conversion
    raw_docs = variables.get('customs_documentation', 0)
    raw_docs_currency = variables.get('customs_documentation_currency', 'USD')
    print(f"[export-debug] customs_documentation: raw={raw_docs}, currency={raw_docs_currency}")

    documentation_usd = to_usd(raw_docs, raw_docs_currency)
    print(f"[export-debug] customs_documentation: after to_usd={documentation_usd}")

    other_costs_usd = to_usd(
        variables.get('brokerage_extra', 0),
        variables.get('brokerage_extra_currency', 'USD')
    )

    # Convert DM Fee from original currency to USD
    dm_fee_value_usd = to_usd(
        variables.get('dm_fee_value', 0),
        variables.get('dm_fee_currency', 'USD')
    )

    # Debug: log key variables
    print(f"[export-debug] logistics: hub={variables.get('logistics_supplier_hub')}, customs={variables.get('logistics_hub_customs')}, client={variables.get('logistics_customs_client')}")
    print(f"[export-debug] time_to_advance_on_receiving={variables.get('time_to_advance_on_receiving')}")
    print(f"[export-debug] brokerage: docs_usd={documentation_usd}, warehousing_usd={warehousing_usd}")

    # Build quote_inputs from variables
    quote_inputs = {
        # Company and terms
        "seller_company": variables.get("seller_company", ""),
        "offer_sale_type": _map_sale_type(variables.get("offer_sale_type", "поставка")),
        "incoterms": variables.get("offer_incoterms", "DDP"),
        "quote_currency": quote.get("currency", "RUB"),
        "delivery_time": variables.get("delivery_time", 30),

        # Advance/Payment percentages
        "advance_to_supplier": variables.get("advance_to_supplier", 0),
        "advance_from_client": variables.get("advance_from_client", 100),
        "time_to_advance": variables.get("time_to_advance", 0),
        "advance_on_loading": variables.get("advance_on_loading", 0),
        "time_to_advance_loading": variables.get("time_to_advance_loading", 0),
        "advance_on_shipping": variables.get("advance_on_shipping", 0),
        "time_to_advance_shipping": variables.get("time_to_advance_shipping", 0),
        "advance_on_customs": variables.get("advance_on_customs", 0),
        "time_to_advance_customs": variables.get("time_to_advance_customs", 0),
        "time_to_advance_on_receiving": variables.get("time_to_advance_on_receiving", 0),

        # Logistics costs (already in USD from invoice aggregation)
        "logistics_supplier_hub": variables.get("logistics_supplier_hub", 0),
        "logistics_hub_customs": variables.get("logistics_hub_customs", 0),
        "logistics_customs_client": variables.get("logistics_customs_client", 0),

        # Brokerage costs (converted from original currency to USD)
        "brokerage_hub": brokerage_hub_usd,
        "brokerage_customs": brokerage_customs_usd,
        "warehousing": warehousing_usd,
        "documentation": documentation_usd,
        "other_costs": other_costs_usd,

        # DM Fee - map to Russian values expected by Excel (value converted to USD)
        "dm_fee_type": _map_dm_fee_type(variables.get("dm_fee_type", "Фикс")),
        "dm_fee_value": dm_fee_value_usd,

        # Admin settings - use same defaults as calculation_mapper.get_default_admin_settings()
        # rate_forex_risk goes to D30, Excel maps it internally to расчет
        "rate_forex_risk": variables.get("rate_forex_risk") if variables.get("rate_forex_risk") is not None else 3,  # Default 3%

        # Internal: exchange rate for USD conversion
        # Get actual USD -> quote_currency rate from currency service
        "_usd_to_quote_rate": _get_usd_to_quote_rate(quote.get("currency", "USD")),
        "_quote_currency": quote.get("currency", "RUB"),
    }

    # Build product_inputs from items
    product_inputs = []
    quote_currency = quote.get("currency", "USD")
    for item in items:
        purchase_currency = item.get("purchase_currency", "USD")
        product_inputs.append({
            "brand": item.get("brand", ""),
            "sku": item.get("product_code", item.get("sku", "")),
            "name": item.get("product_name", ""),
            "quantity": item.get("quantity", 1),
            "weight_in_kg": item.get("weight_kg", 0),
            "currency_of_base_price": purchase_currency,
            "base_price_vat": item.get("purchase_price_original", item.get("base_price_vat", 0)),
            "supplier_country": item.get("supplier_country", ""),
            "supplier_discount": variables.get("supplier_discount", 0),
            # Exchange rate: from purchase currency to quote currency (4 decimal places, CBR)
            "exchange_rate": _get_exchange_rate_to_quote(purchase_currency, quote_currency),
            "customs_code": item.get("customs_code", ""),
            # import_tariff: customs_duty is saved by customs workspace, fallback to import_tariff
            "import_tariff": item.get("customs_duty") or item.get("import_tariff", 0),
            "markup": variables.get("markup", 15),
        })

    # Build api_results by summing from item calculations (source of truth)
    # quote_calculation_summaries may have stale data, so we always calculate from items
    total_purchase = Decimal("0")
    total_logistics_first = Decimal("0")
    total_logistics_last = Decimal("0")
    total_logistics = Decimal("0")
    total_cogs = Decimal("0")
    total_revenue = Decimal("0")  # AK = sales_price_total_no_vat
    total_revenue_with_vat = Decimal("0")  # AL
    total_profit = Decimal("0")  # AF

    for item in items:
        calc = item.get("calc", {})
        total_purchase += Decimal(str(calc.get("S16", 0) or 0))
        total_logistics_first += Decimal(str(calc.get("T16", 0) or 0))
        total_logistics_last += Decimal(str(calc.get("U16", 0) or 0))
        total_logistics += Decimal(str(calc.get("V16", 0) or 0))
        total_cogs += Decimal(str(calc.get("AB16", 0) or 0))
        total_revenue += Decimal(str(calc.get("AK16", 0) or 0))
        total_revenue_with_vat += Decimal(str(calc.get("AL16", 0) or 0))
        total_profit += Decimal(str(calc.get("AF16", 0) or 0))

    api_results = {
        "total_purchase_price": float(total_purchase),
        "total_logistics_first": float(total_logistics_first),
        "total_logistics_last": float(total_logistics_last),
        "total_logistics": float(total_logistics),
        "total_cogs": float(total_cogs),
        "total_revenue": float(total_revenue),
        "total_revenue_with_vat": float(total_revenue_with_vat),
        "total_profit": float(total_profit),

        # Financing (may not be available)
        "evaluated_revenue": float(total_revenue_with_vat),
        "client_advance": 0,
        "total_before_forwarding": 0,
        "supplier_payment": float(total_purchase),
        "supplier_financing_cost": 0,
        "operational_financing_cost": 0,
        "total_financing_cost": 0,
        "credit_sales_amount": 0,
        "credit_sales_fv": 0,
        "credit_sales_interest": 0,

        # Internal
        "_usd_to_quote_rate": _get_usd_to_quote_rate(quote.get("currency", "USD")),
        "_quote_currency": quote.get("currency", "RUB"),
    }

    # Build product_results from item calculations
    product_results = []
    for item in items:
        calc = item.get("calc", {})
        product_results.append({
            "purchase_price_no_vat": calc.get("N16", 0),
            "purchase_price_after_discount": calc.get("P16", 0),
            "purchase_price_per_unit_quote_currency": calc.get("R16", 0),
            "purchase_price_total_quote_currency": calc.get("S16", 0),
            "logistics_first_leg": calc.get("T16", 0),
            "logistics_last_leg": calc.get("U16", 0),
            "logistics_total": calc.get("V16", 0),
            "customs_fee": calc.get("Y16", 0),
            "excise_tax_amount": calc.get("Z16", 0),
            "cogs_per_unit": calc.get("AA16", 0),
            "cogs_per_product": calc.get("AB16", 0),
            "sale_price_per_unit_excl_financial": calc.get("AD16", 0),
            "sale_price_total_excl_financial": calc.get("AE16", 0),
            "profit": calc.get("AF16", 0),
            "dm_fee": calc.get("AG16", 0),
            "forex_reserve": calc.get("AH16", 0),
            "financial_agent_fee": calc.get("AI16", 0),
            "sales_price_per_unit_no_vat": calc.get("AJ16", 0),
            "sales_price_total_no_vat": calc.get("AK16", 0),
            "sales_price_total_with_vat": calc.get("AL16", 0),
            "sales_price_per_unit_with_vat": calc.get("AM16", 0),
            "vat_from_sales": calc.get("AN16", 0),
            "vat_on_import": calc.get("AO16", 0),
            "vat_net_payable": calc.get("AP16", 0),
            "transit_commission": calc.get("AQ16", 0),
            "internal_sale_price_per_unit": calc.get("AX16", 0),
            "internal_sale_price_total": calc.get("AY16", 0),
            "financing_cost_initial": calc.get("BA16", 0),
            "financing_cost_credit": calc.get("BB16", 0),
        })

    # Generate using the main service
    return generate_validation_export(
        quote_inputs, product_inputs, api_results, product_results
    )
