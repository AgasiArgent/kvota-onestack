"""
Validation Excel Export Service

Generates internal QA spreadsheet with auto-highlighting.
Used by sales managers to validate calculations before sending to clients.
"""

from decimal import Decimal
from typing import Dict, Any, List
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.export_data_mapper import ExportData, format_date_russian


# Color definitions
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_validation_excel(data: ExportData) -> bytes:
    """
    Create validation Excel spreadsheet.

    Args:
        data: ExportData from export_data_mapper

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    quote = data.quote
    customer = data.customer
    organization = data.organization
    items = data.items
    variables = data.variables
    calculations = data.calculations

    currency = quote.get("currency", "RUB")
    currency_symbol = {"RUB": "₽", "USD": "$", "EUR": "€", "CNY": "¥"}.get(currency, currency)

    row = 1

    # ==================== HEADER SECTION ====================
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "VALIDATION REPORT"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 2

    # Quote info
    info_data = [
        ("Quote Number:", quote.get("quote_number", "-")),
        ("Customer:", customer.get("company_name", customer.get("name", "-"))),
        ("Date:", format_date_russian()),
        ("Currency:", currency),
        ("Status:", quote.get("status", "draft")),
    ]

    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # ==================== VARIABLES SECTION ====================
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "CALCULATION VARIABLES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = HEADER_FILL
    ws[f'A{row}'].font = HEADER_FONT
    row += 1

    var_data = [
        ("Seller Company:", variables.get("seller_company", "-")),
        ("Incoterms:", variables.get("offer_incoterms", "-")),
        ("Markup:", f"{variables.get('markup', 0)}%"),
        ("Supplier Discount:", f"{variables.get('supplier_discount', 0)}%"),
        ("Delivery Time:", f"{variables.get('delivery_time', 0)} days"),
        ("Client Advance:", f"{variables.get('advance_from_client', 0)}%"),
        ("Exchange Rate:", variables.get("exchange_rate", "1.0")),
    ]

    for label, value in var_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = str(value)
        row += 1

    row += 1

    # Logistics section
    ws[f'A{row}'] = "Logistics (W2-W4):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    logistics_data = [
        ("  Supplier→Hub (W2):", f"{currency_symbol}{variables.get('logistics_supplier_hub', 0):,.2f}"),
        ("  Hub→Customs (W3):", f"{currency_symbol}{variables.get('logistics_hub_customs', 0):,.2f}"),
        ("  Customs→Client (W4):", f"{currency_symbol}{variables.get('logistics_customs_client', 0):,.2f}"),
    ]

    for label, value in logistics_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # Brokerage section
    ws[f'A{row}'] = "Brokerage (W5-W9):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    brokerage_data = [
        ("  Hub Brokerage (W5):", f"{currency_symbol}{variables.get('brokerage_hub', 0):,.2f}"),
        ("  Customs Brokerage (W6):", f"{currency_symbol}{variables.get('brokerage_customs', 0):,.2f}"),
        ("  Warehousing (W7):", f"{currency_symbol}{variables.get('warehousing_at_customs', 0):,.2f}"),
        ("  Documentation (W8):", f"{currency_symbol}{variables.get('customs_documentation', 0):,.2f}"),
        ("  Extra (W9):", f"{currency_symbol}{variables.get('brokerage_extra', 0):,.2f}"),
    ]

    for label, value in brokerage_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1

    row += 2

    # ==================== PRODUCTS TABLE ====================
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = "PRODUCT CALCULATIONS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = HEADER_FILL
    ws[f'A{row}'].font = HEADER_FONT
    row += 1

    # Column headers
    headers = [
        "№", "Product Name", "SKU", "Qty", "Base Price",
        "Purchase (S16)", "Logistics (V16)", "COGS (AB16)",
        "Markup", "Price/Unit (AJ16)", "Total (AK16)",
        "VAT", "Final (AL16)", "Profit (AF16)", "Margin %", "Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    row += 1
    start_data_row = row

    # Product data rows
    for i, item in enumerate(items, 1):
        calc = item.get("calc", {})

        # Calculate margin percentage
        cogs = Decimal(str(calc.get("AB16", 0)))
        profit = Decimal(str(calc.get("AF16", 0)))
        margin_pct = (profit / cogs * 100) if cogs else Decimal("0")

        # Determine status
        markup_var = Decimal(str(variables.get("markup", 15)))
        if margin_pct < Decimal("5"):
            status = "LOW MARGIN"
            status_fill = ERROR_FILL
        elif margin_pct < markup_var:
            status = "BELOW TARGET"
            status_fill = WARNING_FILL
        else:
            status = "OK"
            status_fill = SUCCESS_FILL

        row_data = [
            i,
            item.get("product_name", "")[:40],
            item.get("product_code", "-"),
            item.get("quantity", 1),
            f"{currency_symbol}{item.get('base_price_vat', 0):,.2f}",
            f"{currency_symbol}{calc.get('S16', 0):,.2f}",
            f"{currency_symbol}{calc.get('V16', 0):,.2f}",
            f"{currency_symbol}{calc.get('AB16', 0):,.2f}",
            f"{variables.get('markup', 0)}%",
            f"{currency_symbol}{calc.get('AJ16', 0):,.2f}",
            f"{currency_symbol}{calc.get('AK16', 0):,.2f}",
            f"{currency_symbol}{calc.get('AP16', 0):,.2f}",
            f"{currency_symbol}{calc.get('AL16', 0):,.2f}",
            f"{currency_symbol}{calc.get('AF16', 0):,.2f}",
            f"{margin_pct:.1f}%",
            status,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center' if col in [1, 4, 9, 15, 16] else 'left')

            # Apply status color to last column
            if col == 16:
                cell.fill = status_fill

            # Highlight margin column based on value
            if col == 15:
                cell.fill = status_fill

        row += 1

    row += 1

    # ==================== TOTALS SECTION ====================
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = "TOTALS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = HEADER_FILL
    ws[f'A{row}'].font = HEADER_FONT
    row += 1

    totals_data = [
        ("Total Purchase (S16):", f"{currency_symbol}{calculations.get('total_purchase', 0):,.2f}"),
        ("Total Logistics (V16):", f"{currency_symbol}{calculations.get('total_logistics', 0):,.2f}"),
        ("Total COGS:", f"{currency_symbol}{calculations.get('total_cogs', 0):,.2f}"),
        ("Total Profit:", f"{currency_symbol}{calculations.get('total_profit', 0):,.2f}"),
        ("Total (no VAT):", f"{currency_symbol}{calculations.get('total_no_vat', 0):,.2f}"),
        ("Total VAT:", f"{currency_symbol}{calculations.get('total_vat', 0):,.2f}"),
        ("TOTAL (with VAT):", f"{currency_symbol}{calculations.get('total_with_vat', 0):,.2f}"),
    ]

    for label, value in totals_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if "TOTAL (with VAT)" in label:
            ws[f'B{row}'].font = Font(bold=True, size=12)
            ws[f'B{row}'].fill = SUCCESS_FILL
        row += 1

    # ==================== COLUMN WIDTHS ====================
    column_widths = [5, 35, 15, 8, 12, 14, 14, 14, 10, 14, 14, 12, 14, 14, 10, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
