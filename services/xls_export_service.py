"""
Invoice XLS Export Service

Generates XLSX file from an invoice and its per-invoice positions
(``kvota.invoice_items``) using openpyxl. Language parameter controls column
headers and item name field.

Phase 5d (Task 3): reads from ``invoice_items`` (not ``quote_items``) with an
embedded PostgREST join through ``invoice_item_coverage`` to pull
quote-level metadata (requested SKU, manufacturer product name, EN name) and
to populate the "Покрывает" / "Covers" column for merged rows.

Called by POST /api/invoices/{id}/download-xls.
"""

import logging
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from services.database import get_supabase

logger = logging.getLogger(__name__)


# --- Column definitions ---
#
# Fields prefixed with "_" are computed by ``_get_cell_value`` from the row
# and its embedded coverage rows. Other field names map directly to
# ``invoice_items`` columns (see migration 281) or to quote_items columns
# accessed via coverage (see ``_get_quote_item_field``).

COLUMNS_RU: list[tuple[str, str]] = [
    ("Бренд", "brand"),
    ("Арт. запрошенный", "_idn_sku"),
    ("Арт. производителя", "supplier_sku"),
    ("Наименование производителя", "_manufacturer_product_name"),
    ("Наименование", "_item_name"),
    ("Кол-во", "quantity"),
    ("Мин. заказ", "minimum_order_quantity"),
    ("Цена", "purchase_price_original"),
    ("Срок (к.д.)", "production_time_days"),
    ("Вес (кг)", "weight_in_kg"),
    ("Размеры (В×Ш×Д мм)", "_dimensions"),
    ("Примечание", "supplier_notes"),
    ("Покрывает", "_covers"),
]

COLUMNS_EN: list[tuple[str, str]] = [
    ("Brand", "brand"),
    ("Requested SKU", "_idn_sku"),
    ("Manufacturer SKU", "supplier_sku"),
    ("Manufacturer Name", "_manufacturer_product_name"),
    ("Item Name", "_item_name"),
    ("Quantity", "quantity"),
    ("MOQ", "minimum_order_quantity"),
    ("Price", "purchase_price_original"),
    ("Lead Time (days)", "production_time_days"),
    ("Weight (kg)", "weight_in_kg"),
    ("Dimensions (HxWxL mm)", "_dimensions"),
    ("Notes", "supplier_notes"),
    ("Covers", "_covers"),
]

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def _get_columns(language: str) -> list[tuple[str, str]]:
    """Return column definitions for the given language."""
    if language == "en":
        return COLUMNS_EN
    return COLUMNS_RU


def _format_dimensions(item: dict[str, Any]) -> str | None:
    """Format dimensions as HxWxL string, or None if all dimensions are missing."""
    h = item.get("dimension_height_mm")
    w = item.get("dimension_width_mm")
    l = item.get("dimension_length_mm")
    if h is None and w is None and l is None:
        return None
    return f"{h or 0}×{w or 0}×{l or 0}"


def _coverage_rows(item: dict[str, Any]) -> list[dict[str, Any]]:
    """Return the embedded coverage list (possibly empty)."""
    return item.get("invoice_item_coverage") or []


def _get_quote_item_field(item: dict[str, Any], field: str) -> Any:
    """Read a field from the item's primary quote_item via coverage.

    For 1:1 and split (single coverage row), this is unambiguous. For merge
    (multiple coverage rows), the first coverage row's quote_item provides
    the value — the full merge membership is rendered separately in the
    "Покрывает" column.

    Returns None when coverage is missing or the linked quote_items row
    is absent (unusual — indicates data inconsistency).
    """
    rows = _coverage_rows(item)
    if not rows:
        return None
    qi = rows[0].get("quote_items") or {}
    return qi.get(field)


def _get_covers(item: dict[str, Any]) -> str | None:
    """Build the "Покрывает" column value.

    Returns None for 1:1 / split rows (single coverage). For merged rows
    (two or more coverage entries), returns a comma-joined list of the
    source quote_items' product_names in the order returned by PostgREST.
    """
    rows = _coverage_rows(item)
    if len(rows) < 2:
        return None
    names = []
    for cov in rows:
        qi = cov.get("quote_items") or {}
        name = qi.get("product_name")
        if name:
            names.append(name)
    return ", ".join(names) if names else None


def _get_item_name(item: dict[str, Any], language: str) -> str | None:
    """Get item name based on language.

    For EN, prefer the quote_item's ``name_en`` (populated by sales),
    falling back to the invoice_item's ``product_name`` when absent. For
    RU, always use the invoice_item's ``product_name`` — the supplier's
    version, which is what the supplier expects to see in their KP.
    """
    if language == "en":
        name_en = _get_quote_item_field(item, "name_en")
        if name_en:
            return name_en
    return item.get("product_name")


def _get_cell_value(item: dict[str, Any], field_key: str, language: str) -> Any:
    """Extract cell value from an item for the given field key."""
    if field_key == "_item_name":
        return _get_item_name(item, language)
    if field_key == "_dimensions":
        return _format_dimensions(item)
    if field_key == "_idn_sku":
        return _get_quote_item_field(item, "idn_sku")
    if field_key == "_manufacturer_product_name":
        return _get_quote_item_field(item, "manufacturer_product_name")
    if field_key == "_covers":
        return _get_covers(item)
    return item.get(field_key)


def _fetch_invoice(invoice_id: str) -> dict[str, Any]:
    """Fetch invoice with supplier name."""
    supabase = get_supabase()
    result = supabase.table("invoices").select(
        "*, suppliers!supplier_id(id, name)"
    ).eq("id", invoice_id).execute()
    data = result.data or []
    if not data:
        raise ValueError(f"Invoice not found: {invoice_id}")
    return data[0]


def _fetch_invoice_items(invoice_id: str) -> list[dict[str, Any]]:
    """Fetch invoice_items rows with coverage + source quote_items metadata.

    Single PostgREST query: pulls every ``invoice_items`` row for the
    invoice and embeds the coverage rows, each with the linked
    quote_items fields needed for display (requested SKU, manufacturer
    name, EN name, and the quote_item's product_name for the "Покрывает"
    column).
    """
    supabase = get_supabase()
    result = (
        supabase.table("invoice_items")
        .select(
            "*, invoice_item_coverage(quote_item_id, ratio, "
            "quote_items(product_name, quantity, idn_sku, "
            "manufacturer_product_name, name_en))"
        )
        .eq("invoice_id", invoice_id)
        .order("position")
        .execute()
    )
    return result.data or []


def generate_invoice_xls(invoice_id: str, language: str = "ru") -> bytes:
    """Generate XLS from invoice + its invoice_items positions.

    Path: called by POST /api/invoices/{id}/download-xls
    Params:
        invoice_id: UUID of the invoice
        language: 'ru' or 'en' -- controls column headers and item name field
    Returns:
        bytes -- the XLSX file content
    """
    invoice = _fetch_invoice(invoice_id)
    items = _fetch_invoice_items(invoice_id)
    columns = _get_columns(language)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Items"

    # Write header row
    for col_idx, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Write item rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, (_, field_key) in enumerate(columns, 1):
            value = _get_cell_value(item, field_key, language)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col_idx, (header, _) in enumerate(columns, 1):
        max_len = len(str(header))
        for row_idx in range(2, len(items) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
