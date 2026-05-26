"""
Invoice XLS Import Service (Testing 2 row 70).

Reverse of services.xls_export_service: clients download the КПП as XLS via
«Скачать XLS», edit it offline, and re-upload through «Загрузить XLS». The
service parses the file, matches rows back to ``kvota.invoice_items`` by
``product_code`` (the customer-side article from МОП, joined via
``invoice_item_coverage`` → ``quote_items``), and updates the supplier-side
fields on each matched row.

Testing 2 row 88: the "Арт. запрошенный" column in the template now carries
``quote_items.product_code`` (not ``idn_sku``) — suppliers identify parts by
article, so the round-trip match key follows the visible column.

Edge cases (docs/plans/2026-05-25-batch-24c-decisions.md, Q5):
  - new article in XLS (not in КПП)      → skip + return in ``skipped`` list
  - article in КПП missing from XLS      → preserve existing values (no-op)
  - duplicate article in XLS             → raise ``DuplicateArticlesError``
  - all editable fields land on UPDATE   → price, qty, MOQ, weight, dims, …

Called by POST /api/invoices/{id}/import-xls (api.invoices.import_invoice_xls).
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from services.database import get_supabase

logger = logging.getLogger(__name__)


# Index of each editable column in the «Скачать XLS» template. The order
# MUST match services.xls_export_service.COLUMNS_RU. Header row is row 1;
# data rows start at row 2. The "Покрывает" column (last) is read-only
# metadata generated on export and is intentionally ignored on import.
# Columns 2 ("Арт. запрошенный" = product_code) and 7 ("Ед. изм." = unit)
# come from the customer side (quote_items) and are not updated on import.
_COL_BRAND = 1
_COL_PRODUCT_CODE = 2
_COL_SUPPLIER_SKU = 3
_COL_MANUFACTURER_NAME = 4  # noqa: F841 — quote_items-side, not imported here
_COL_PRODUCT_NAME = 5
_COL_QUANTITY = 6
_COL_UNIT = 7  # noqa: F841 — quote_items.unit, read-only on import
_COL_MOQ = 8
_COL_PRICE = 9
_COL_LEAD_TIME = 10
_COL_WEIGHT = 11
_COL_DIMENSIONS = 12
_COL_NOTES = 13


class DuplicateArticlesError(Exception):
    """Raised when the uploaded XLS contains the same article more than once.

    The endpoint converts this to a 400 with a ``DUPLICATES`` error code
    listing the offending articles (matches «Дубликаты артикулов: [list]»
    toast on the frontend).
    """

    def __init__(self, duplicates: list[str]) -> None:
        super().__init__(f"Duplicate articles in XLS: {duplicates}")
        self.duplicates = duplicates


def _parse_int(value: Any) -> int | None:
    """Coerce to int — empty / None / unparseable returns None."""
    if value is None or value == "":
        return None
    try:
        return int(float(value))  # accept "5" and 5.0 alike
    except (TypeError, ValueError):
        return None


def _parse_float(value: Any) -> float | None:
    """Coerce to float — empty / None / unparseable returns None."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_text(value: Any) -> str | None:
    """Coerce to non-empty string — empty / None returns None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _parse_dimensions(value: Any) -> tuple[int | None, int | None, int | None]:
    """Parse "H×W×L" / "HxWxL" / "H*W*L" / "H W L" into a 3-tuple of ints.

    Mirrors ``parseDimensions`` in procurement-handsontable.tsx so users
    can paste between the two surfaces without surprises.
    """
    if value is None or value == "":
        return (None, None, None)
    raw = str(value).strip()
    if not raw:
        return (None, None, None)
    # Accept ×, x, X, *, whitespace as separators
    parts = []
    current = ""
    for ch in raw:
        if ch in ("×", "x", "X", "*") or ch.isspace():
            if current:
                parts.append(current)
                current = ""
        else:
            current += ch
    if current:
        parts.append(current)

    def _to_int(s: str) -> int | None:
        try:
            return int(float(s))
        except (TypeError, ValueError):
            return None

    h = _to_int(parts[0]) if len(parts) >= 1 else None
    w = _to_int(parts[1]) if len(parts) >= 2 else None
    length = _to_int(parts[2]) if len(parts) >= 3 else None
    return (h, w, length)


def _build_update_payload(row: dict[str, Any]) -> dict[str, Any]:
    """Map a parsed XLS row to the ``invoice_items`` UPDATE payload.

    All template fields are included so editing any cell takes effect.
    Empty cells write ``None`` — this is intentional: a blank in XLS means
    "clear the field" (same semantic as deleting a cell in the inline
    handsontable). Identifying the row is the caller's job.
    """
    h, w, length = _parse_dimensions(row.get("dimensions"))
    return {
        "brand": _parse_text(row.get("brand")),
        "supplier_sku": _parse_text(row.get("supplier_sku")),
        "product_name": _parse_text(row.get("product_name")) or "",
        "quantity": _parse_int(row.get("quantity")),
        "minimum_order_quantity": _parse_int(row.get("minimum_order_quantity")),
        "purchase_price_original": _parse_float(row.get("purchase_price_original")),
        "production_time_days": _parse_int(row.get("production_time_days")),
        "weight_in_kg": _parse_float(row.get("weight_in_kg")),
        "dimension_height_mm": h,
        "dimension_width_mm": w,
        "dimension_length_mm": length,
        "supplier_notes": _parse_text(row.get("supplier_notes")),
    }


def _parse_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    """Read the template xlsx and return one dict per data row.

    Rows with an empty `product_code` cell are dropped silently — pasting a
    few blank rows at the bottom of the sheet is common and shouldn't be
    surfaced as "skipped".
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows: list[dict[str, Any]] = []
    # Start at row 2 — row 1 is the header (we don't validate it; the column
    # ORDER is what the contract guarantees, same as «Скачать XLS»).
    # Template has 13 data columns (last is read-only "Покрывает").
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if excel_row is None:
            continue
        # Pad the tuple to the expected length so missing trailing cells
        # don't IndexError.
        cells = list(excel_row) + [None] * (13 - len(excel_row))
        product_code = _parse_text(cells[_COL_PRODUCT_CODE - 1])
        if not product_code:
            # Blank-article row → ignore silently
            continue
        rows.append(
            {
                "brand": cells[_COL_BRAND - 1],
                "product_code": product_code,
                "supplier_sku": cells[_COL_SUPPLIER_SKU - 1],
                "product_name": cells[_COL_PRODUCT_NAME - 1],
                "quantity": cells[_COL_QUANTITY - 1],
                "minimum_order_quantity": cells[_COL_MOQ - 1],
                "purchase_price_original": cells[_COL_PRICE - 1],
                "production_time_days": cells[_COL_LEAD_TIME - 1],
                "weight_in_kg": cells[_COL_WEIGHT - 1],
                "dimensions": cells[_COL_DIMENSIONS - 1],
                "supplier_notes": cells[_COL_NOTES - 1],
            }
        )
    return rows


def _fetch_kpp_lookup(invoice_id: str) -> dict[str, str]:
    """Build {product_code → invoice_item_id} for every row in the КПП.

    Single PostgREST query — same shape used by xls_export_service so the
    coverage embed comes back populated. Each invoice_item joins to one or
    more quote_items via ``invoice_item_coverage``; we index by the FIRST
    coverage's ``product_code`` since merged rows (multiple coverages) share
    a common identifier on the supplier КП (the template's "Арт. запрошенный"
    column likewise reads ``rows[0]``).
    """
    sb = get_supabase()
    result = (
        sb.table("invoice_items")
        .select(
            "id, invoice_item_coverage(quote_item_id, ratio, "
            "quote_items(product_code))"
        )
        .eq("invoice_id", invoice_id)
        .order("position")
        .execute()
    )
    items = getattr(result, "data", None) or []
    lookup: dict[str, str] = {}
    for item in items:
        coverage = item.get("invoice_item_coverage") or []
        if not coverage:
            continue
        first_qi = coverage[0].get("quote_items") or {}
        product_code = first_qi.get("product_code")
        if not product_code:
            continue
        # Stringify defensively + trim: trailing whitespace from data-entry
        # should not block a match.
        key = str(product_code).strip()
        if key:
            lookup[key] = str(item["id"])
    return lookup


def import_invoice_xls(invoice_id: str, file_bytes: bytes) -> dict[str, Any]:
    """Parse + apply an XLS import to a КПП invoice.

    Params:
        invoice_id: UUID of the target invoice (КПП).
        file_bytes: Raw xlsx bytes from the multipart upload.
    Returns:
        ``{"updated": N, "skipped": [list-of-articles], "total_in_file": M}``
        where ``skipped`` lists articles present in the file but not in the
        КПП. Articles present in the КПП but absent from the file are NOT
        included (silently preserved).
    Raises:
        DuplicateArticlesError: if the file contains the same article twice.
    Side Effects:
        ``UPDATE kvota.invoice_items SET ...`` for every matched row.
    """
    parsed = _parse_xlsx(file_bytes)
    total_in_file = len(parsed)

    # --- Duplicate guard (locked decision: reject the whole upload) ----------
    seen: set[str] = set()
    duplicates: list[str] = []
    for row in parsed:
        product_code = row["product_code"]
        if product_code in seen and product_code not in duplicates:
            duplicates.append(product_code)
        seen.add(product_code)
    if duplicates:
        raise DuplicateArticlesError(duplicates)

    # --- Match-key lookup ----------------------------------------------------
    lookup = _fetch_kpp_lookup(invoice_id)

    # --- Apply updates -------------------------------------------------------
    sb = get_supabase()
    updated = 0
    skipped: list[str] = []
    for row in parsed:
        product_code = row["product_code"]
        invoice_item_id = lookup.get(product_code)
        if invoice_item_id is None:
            skipped.append(product_code)
            continue

        payload = _build_update_payload(row)
        try:
            sb.table("invoice_items").update(payload).eq("id", invoice_item_id).execute()
            updated += 1
        except Exception as exc:
            # One bad row should not poison the whole import. Log + continue.
            logger.error(
                "XLS import: failed to update invoice_item %s for article %s: %s",
                invoice_item_id,
                product_code,
                exc,
            )
            # Treat as skipped so the caller sees the article name surface in
            # the toast — the user can then retry from a fresh download.
            skipped.append(product_code)

    return {
        "updated": updated,
        "skipped": skipped,
        "total_in_file": total_in_file,
    }
