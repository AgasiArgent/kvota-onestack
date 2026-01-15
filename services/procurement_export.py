"""
Procurement Excel Export Service

Generates Excel spreadsheet with quote items for procurement managers
to send to suppliers for pricing.

Feature #36: Скачивание списка для оценки
"""

from typing import Dict, Any, List
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Color definitions
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_procurement_excel(
    quote: Dict[str, Any],
    items: List[Dict[str, Any]],
    brands: List[str],
    customer_name: str = ""
) -> bytes:
    """
    Create procurement Excel spreadsheet for sending to suppliers.

    This export contains:
    - Quote info header
    - Items grouped by brand
    - Columns for supplier to fill in: price, availability, lead time

    Args:
        quote: Quote data dict
        items: List of quote items (already filtered for user's brands)
        brands: List of user's assigned brands
        customer_name: Customer company name for header

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Запрос на оценку"

    quote_number = quote.get("idn_quote", quote.get("id", "")[:8])
    currency = quote.get("currency", "RUB")
    currency_symbol = {"RUB": "₽", "USD": "$", "EUR": "€", "CNY": "¥"}.get(currency, currency)

    row = 1

    # ==================== HEADER SECTION ====================
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "ЗАПРОС НА ОЦЕНКУ ПОЗИЦИЙ"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2

    # Quote info
    info_data = [
        ("Номер КП:", quote_number),
        ("Клиент:", customer_name or "—"),
        ("Дата запроса:", datetime.now().strftime("%d.%m.%Y")),
        ("Валюта:", currency),
    ]

    for label, value in info_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # Instructions
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "Пожалуйста, заполните колонки: Цена поставщика, Наличие, Срок производства"
    ws[f'A{row}'].font = Font(italic=True, color="666666")
    row += 2

    # ==================== ITEMS TABLE ====================
    # Column headers
    headers = [
        "№",
        "Бренд",
        "Артикул",
        "Наименование",
        "Кол-во",
        "Ед.",
        "Цена поставщика",  # For supplier to fill
        "Наличие",          # For supplier to fill
        "Срок пр-ва, дн.",  # For supplier to fill
        "Примечания"        # For supplier to fill
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    row += 1

    # Group items by brand
    items_by_brand = {}
    for item in items:
        brand = item.get("brand", "Прочее")
        if brand not in items_by_brand:
            items_by_brand[brand] = []
        items_by_brand[brand].append(item)

    # Sort brands alphabetically
    sorted_brands = sorted(items_by_brand.keys())

    item_num = 1
    for brand in sorted_brands:
        brand_items = items_by_brand[brand]

        # Brand subheader
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f"▸ {brand} ({len(brand_items)} позиций)"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = SUBHEADER_FILL
        ws[f'A{row}'].border = THIN_BORDER
        row += 1

        # Items in this brand
        for item in brand_items:
            row_data = [
                item_num,
                brand,
                item.get("product_code", ""),
                item.get("name", item.get("product_name", "")),
                item.get("quantity", 1),
                item.get("unit", "шт."),
                "",  # Price - for supplier to fill
                "",  # Availability - for supplier to fill
                "",  # Lead time - for supplier to fill
                "",  # Notes - for supplier to fill
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = THIN_BORDER

                # Alignment
                if col in [1, 5, 7, 9]:  # Numbers
                    cell.alignment = Alignment(horizontal='center')
                elif col in [6, 8]:  # Short text
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')

                # Highlight columns for supplier to fill
                if col in [7, 8, 9, 10]:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            item_num += 1
            row += 1

        row += 1  # Empty row between brands

    # ==================== FOOTER ====================
    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f"Всего позиций: {len(items)}"
    ws[f'A{row}'].font = Font(bold=True)
    row += 2

    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "Заполненный файл отправьте обратно менеджеру по закупкам."
    ws[f'A{row}'].font = Font(italic=True, color="666666")

    # ==================== COLUMN WIDTHS ====================
    column_widths = [5, 15, 20, 40, 8, 8, 15, 12, 12, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A9'  # Freeze above the data table header

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def create_procurement_simple_list(
    items: List[Dict[str, Any]],
    brand_filter: str = None
) -> bytes:
    """
    Create a simple list of items for quick copy-paste to supplier.

    Minimal columns: brand, code, name, quantity

    Args:
        items: List of quote items
        brand_filter: Optional - filter by specific brand

    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Список позиций"

    # Filter by brand if specified
    if brand_filter:
        items = [i for i in items if i.get("brand", "").lower() == brand_filter.lower()]

    row = 1

    # Simple headers
    headers = ["Бренд", "Артикул", "Наименование", "Кол-во"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row += 1

    # Items
    for item in items:
        row_data = [
            item.get("brand", ""),
            item.get("product_code", ""),
            item.get("name", item.get("product_name", "")),
            item.get("quantity", 1),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = THIN_BORDER

        row += 1

    # Column widths
    column_widths = [15, 20, 50, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
